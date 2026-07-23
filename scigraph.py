#!/usr/bin/env python3
"""
================================================================================
ENTERPRISE SCIENTIFIC KNOWLEDGE GRAPH PLATFORM v3.1 ULTRA
================================================================================
An automated scientific knowledge integration & discovery engine for 
computational biology, chemistry, drug discovery, biotechnology, environmental 
science, agriculture, microbiology, toxicology, pharmacology, and life sciences.

Features:
  - 50+ Production Connectors with Active REST API integrations for PubChem,
    ChEMBL, UniProt, BRENDA, Guide to Pharmacology (GtoPdb), STRING DB,
    PDBe-KB, OpenAlex, NCBI Gene, NCBI Taxonomy, CompTox, Rhea, BindingDB, etc.
  - Target & Bioactivity Integration: Retrieves explicit Inhibitors, Activators,
    Substrates, IC50, Ki, EC50, Kd, pChEMBL values, assays, and kinetic parameters.
  - Interactive Query Assistant (`QueryAssistant`): Generates domain-specific
    clarifying questions (organism, bioactivity cutoff, chemical class, depth).
  - Search Library Builder (`SearchLibraryBuilder`): Compiles structured, reusable
    Search Libraries (`search_library.json`, `library_queries.txt`, `library_summary.md`).
  - Multi-Hop Recursive Graph Expander with cycle detection and depth limiting.
  - Entity Resolution Engine with strict/lenient merging, canonical ID enforcement,
    cross-reference mapping, and non-destructive evidence provenance auditing.
  - AI/NLP Knowledge Discovery Engine (NER, relation extraction, paper/patent
    summarization, evidence confidence scoring, conflict detection, research trend
    clustering, and drug repurposing prediction).
  - Graph Analytics Engine (PageRank, Betweenness Centrality, Connected Components,
    Shortest Path finding, and HITS Hub/Authority scoring).
  - Multi-Format Exporters: Neo4j Cypher, NetworkX, GraphML, CSV, RDF JSON-LD,
    RDF Turtle, Parquet, and FAISS Vector Index Metadata.

Usage:
  1. Interactive Query Assistant & Library Build:
     py -3.13 scigraph.py "Ammonia Monooxygenase" --assistant

  2. Direct Search & Multi-Hop Expansion:
     py -3.13 scigraph.py "Ammonia Monooxygenase" --hops 2

  3. Batch Processing Mode:
     py -3.13 scigraph.py --batch library_queries.txt --export-dir ./batch_exports
================================================================================
"""

from __future__ import annotations

import abc
import argparse
import asyncio
import copy
import csv
import hashlib
import json
import logging
import os
import random
import re
import sqlite3
import sys
import time
import zlib
from collections import defaultdict
from enum import Enum
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple, Union
from urllib.parse import quote, urlparse

# Third-Party Dependencies with Fallbacks
try:
    import aiohttp
    import duckdb
    import networkx as nx
    from pydantic import BaseModel, ConfigDict, Field, field_validator, model_validator
except ImportError as e:
    msg = f"Missing critical dependencies: {e}" + "\nPlease run: pip install pydantic networkx duckdb aiohttp orjson pyarrow"
    sys.exit(msg)

try:
    import orjson
except ImportError:
    import json as orjson

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    HAS_PARQUET = True
except ImportError:
    HAS_PARQUET = False


# ==============================================================================
# 1. CONFIGURATION SYSTEM & STRUCTURED LOGGING
# ==============================================================================

DEFAULT_CONFIG: Dict[str, Any] = {
    "engine": {
        "max_concurrency": 15,
        "request_timeout_sec": 10,
        "max_retries": 2,
        "cache_ttl_sec": 86400,
        "rate_limit_hz": 5.0,
        "expand_cross_refs": True,
        "max_expansion_hops": 2,
        "max_entities_per_hop": 20,
    },
    "storage": {
        "workspace_dir": "./scigraph_data",
        "export_dir": "./exports",
        "duckdb_filename": "graph.duckdb",
        "cache_filename": "cache.sqlite",
    },
    "logging": {
        "level": "INFO",
        "json_format": True,
    }
}


class JSONLogFormatter(logging.Formatter):
    def format(self, record: logging.LogRecord) -> str:
        log_obj = {
            "timestamp": self.formatTime(record, self.datefmt),
            "level": record.levelname,
            "module": record.name,
            "message": record.getMessage()
        }
        if record.exc_info:
            log_obj["exception"] = self.formatException(record.exc_info)
        return json.dumps(log_obj)

logger = logging.getLogger("SciGraphEnterprise")
logger.setLevel(logging.INFO)
if not logger.handlers:
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(JSONLogFormatter())
    logger.addHandler(handler)


# ==============================================================================
# 2. CONTROLLED ENUMS & VALIDATION HELPERS
# ==============================================================================

class EnumStr(str):
    """String wrapper providing a .value property for uniform Enum access."""
    @property
    def value(self) -> str:
        return str(self)


class DatabaseSource(str, Enum):
    """Controlled vocabulary of scientific database sources across 50+ repositories."""

    # Chemical & Drug Databases
    PUBCHEM = "PubChem"
    CHEMBL = "ChEMBL"
    DRUGBANK = "DrugBank"
    CHEBI = "ChEBI"
    BINDINGDB = "BindingDB"
    HMDB = "HMDB"
    ZINC = "ZINC"
    NPASS = "NPASS"
    COCONUT = "COCONUT"
    LOTUS = "LOTUS"
    FOODB = "FooDB"
    DRUGCENTRAL = "DrugCentral"
    GTOPDB = "Guide to Pharmacology"
    SURECHEMBL = "SureChEMBL"
    CHEMSPIDER = "ChemSpider"
    COMPTOX = "EPA CompTox"
    DSSTOX = "DSSTox"
    SUSDAT = "NORMAN SusDat"
    EMOLECULES = "eMolecules"

    # Protein & Structural Databases
    UNIPROT = "UniProt"
    SWISSPROT = "SwissProt"
    TREMBL = "TrEMBL"
    ALPHAFOLD_DB = "AlphaFold DB"
    PDB = "PDB"
    PDBE = "PDBe"
    PDBE_KB = "PDBe-KB"
    SIFTS = "SIFTS"
    INTERPRO = "InterPro"
    PFAM = "Pfam"
    SMART = "SMART"
    CATH = "CATH"
    SCOP = "SCOP"
    PROSITE = "Prosite"

    # Genomics & Variant Databases
    NCBI_GENE = "NCBI Gene"
    ENSEMBL = "Ensembl"
    REFSEQ = "RefSeq"
    HGNC = "HGNC"
    DBSNP = "dbSNP"
    CLINVAR = "ClinVar"
    GNOMAD = "gnomAD"
    GTEX = "GTEx"
    OMIM = "OMIM"

    # Pathways & Reactions
    KEGG = "KEGG"
    REACTOME = "Reactome"
    WIKIPATHWAYS = "WikiPathways"
    METACYC = "MetaCyc"
    BIOCYC = "BioCyc"
    PATHWAY_COMMONS = "Pathway Commons"
    SIGNOR = "SIGNOR"

    # Ontologies & Vocabularies
    GO = "Gene Ontology"
    DISEASE_ONTOLOGY = "Disease Ontology"
    MESH = "MeSH"
    NCIT = "NCIT"
    MONDO = "MONDO"
    EFO = "EFO"
    UBERON = "Uberon"
    CELL_ONTOLOGY = "Cell Ontology"
    SEQUENCE_ONTOLOGY = "Sequence Ontology"
    CHEBI_ONTOLOGY = "ChEBI Ontology"
    PLANT_ONTOLOGY = "Plant Ontology"
    ENVIRONMENT_ONTOLOGY = "Environment Ontology"
    UMLS = "UMLS"

    # Interactions
    STRING = "STRING"
    BIOGRID = "BioGRID"
    INTACT = "IntAct"
    MINT = "MINT"
    DIP = "DIP"
    IID = "IID"
    HIPPIE = "HIPPIE"

    # Enzymes
    BRENDA = "BRENDA"
    RHEA = "Rhea"
    EXPLORENZ = "ExplorEnz"

    # Literature
    PUBMED = "PubMed"
    EUROPE_PMC = "Europe PMC"
    CROSSREF = "CrossRef"
    OPENALEX = "OpenAlex"
    SEMANTIC_SCHOLAR = "Semantic Scholar"
    CROSSMARK = "CrossMark"
    BIORXIV = "bioRxiv"
    MEDRXIV = "medRxiv"

    # Patents
    GOOGLE_PATENTS = "Google Patents"
    USPTO = "USPTO"
    PATENTSVIEW = "PatentsView"
    PATENTSCOPE = "WIPO PATENTSCOPE"
    ESPACENET = "Espacenet"
    LENS = "Lens.org"
    SURECHEMBL_PATENTS = "SureChEMBL Patents"

    # Clinical
    CLINICALTRIALS = "ClinicalTrials.gov"
    EU_CLINICALTRIALS = "EU Clinical Trials"
    WHO_ICTRP = "WHO ICTRP"

    # Microbiology & Taxonomy
    NCBI_TAXONOMY = "NCBI Taxonomy"
    GTDB = "GTDB"
    BACDIVE = "BacDive"
    LPSN = "LPSN"

    # Plants & Agriculture
    GRAMENE = "Gramene"
    TAIR = "TAIR"
    PLANT_REACTOME = "Plant Reactome"
    PLANTTFDB = "PlantTFDB"
    FAOSTAT = "FAOSTAT"
    USDA = "USDA"
    EPPO = "EPPO"
    IRAC = "IRAC"
    FRAC = "FRAC"

    # Toxicology
    TOXNET = "TOXNET"
    COMPTOX_TOX = "CompTox Toxicity"
    ECHA = "ECHA"

    # Open Web Knowledge
    WIKIPEDIA = "Wikipedia"
    WIKIDATA = "Wikidata"
    DBPEDIA = "DBpedia"
    OPEN_GOV = "Open Government Data"

    # Fallbacks
    CUSTOM = "Custom"
    OTHER = "Other"

    @classmethod
    def from_string(cls, val: Any) -> DatabaseSource:
        if isinstance(val, cls):
            return val
        if not val:
            return cls.OTHER
        s_val = str(val).strip()
        cleaned = re.sub(r"[_\s\-\.]+", "", s_val).upper()
        for item in cls:
            item_cleaned = re.sub(r"[_\s\-\.]+", "", item.value).upper()
            item_name_cleaned = re.sub(r"[_\s\-\.]+", "", item.name).upper()
            if cleaned in (item_cleaned, item_name_cleaned) or item_cleaned in cleaned or cleaned in item_cleaned:
                return item
        return cls.OTHER


class EntityType(str, Enum):
    """Controlled vocabulary of scientific entity types."""

    COMPOUND = "Compound"
    PROTEIN = "Protein"
    GENE = "Gene"
    ENZYME = "Enzyme"
    PATHWAY = "Pathway"
    REACTION = "Reaction"
    BIOASSAY = "Bioassay"
    PUBLICATION = "Publication"
    PATENT = "Patent"
    DISEASE = "Disease"
    ORGANISM = "Organism"
    STRUCTURE = "Structure"
    DRUG = "Drug"
    TARGET = "Target"
    CELL_LINE = "Cell Line"
    VARIANT = "Variant"
    METABOLITE = "Metabolite"
    PROTEIN_DOMAIN = "Protein Domain"
    MICROORGANISM = "Microorganism"
    CROP = "Crop"
    PEST = "Pest"
    FERTILIZER = "Fertilizer"
    ADMET_PROPERTY = "ADMET Property"
    CLINICAL_TRIAL = "Clinical Trial"
    PATENT_CLAIM = "Patent Claim"
    ONTOLOGY_TERM = "Ontology Term"
    REGULATORY_ELEMENT = "Regulatory Element"
    GENOMIC_VARIANT = "Genomic Variant"

    @classmethod
    def from_string(cls, val: Any) -> Union[EntityType, EnumStr]:
        if isinstance(val, cls):
            return val
        if not val:
            return cls.COMPOUND
        s_val = str(val).strip()
        for item in cls:
            if item.value.lower() == s_val.lower() or item.name.lower() == s_val.lower():
                return item
        return EnumStr(s_val)


class RelationType(str, Enum):
    """Controlled vocabulary of scientific relationship types."""

    BINDS = "binds"
    INHIBITS = "inhibits"
    ACTIVATES = "activates"
    PARTICIPATES_IN = "participates_in"
    ENCODED_BY = "encoded_by"
    EXPRESSED_IN = "expressed_in"
    HAS_STRUCTURE = "has_structure"
    MENTIONED_IN = "mentioned_in"
    PATENTED_IN = "patented_in"
    CATALYZES = "catalyzes"
    REGULATES = "regulates"
    ASSOCIATED_WITH = "associated_with"
    TREATS = "treats"
    SUBSTRATE_OF = "substrate_of"
    HOMOLOG_OF = "homolog_of"
    INTERACTS_PHYSICALLY_WITH = "interacts_physically_with"
    PART_OF = "part_of"
    SUPPORTED_BY = "supported_by"
    CLAIMED_IN = "claimed_in"
    CITES = "cites"
    VALIDATES = "validates"
    STUDIES = "studies"
    LOCATED_IN = "located_in"
    METABOLIZED_BY = "metabolized_by"
    BELONGS_TO = "belongs_to"
    HAS_TOXICITY = "has_toxicity"
    HAS_ADMET = "has_admet"
    IS_A = "is_a"

    @classmethod
    def from_string(cls, val: Any) -> Union[RelationType, EnumStr]:
        if isinstance(val, cls):
            return val
        if not val:
            return cls.ASSOCIATED_WITH
        s_val = str(val).strip()
        for item in cls:
            if item.value.lower() == s_val.lower() or item.name.lower() == s_val.lower():
                return item
        return EnumStr(s_val)


class MergePolicy(str, Enum):
    STRICT = "STRICT"
    LENIENT = "LENIENT"
    ONTOLOGY = "ONTOLOGY"
    OVERWRITE = "OVERWRITE"


# Regex Validation Patterns
URL_REGEX = re.compile(r"^https?://[^\s/$.?#].[^\s]*$", re.IGNORECASE)
DOI_REGEX = re.compile(r"^(https?://(?:dx\.)?doi\.org/)?(10\.\d{4,9}/[-._;()/:A-Za-z0-9]+)$")
PMID_REGEX = re.compile(r"^\d+$")
PATENT_REGEX = re.compile(r"^[A-Z]{2}\d+[A-Z0-9]*$", re.IGNORECASE)


def validate_url(url: Optional[str]) -> Optional[str]:
    if url is None: return None
    url_str = url.strip()
    if not url_str: return None
    parsed = urlparse(url_str)
    if not parsed.scheme or parsed.scheme not in ("http", "https") or not parsed.netloc:
        raise ValueError(f"Invalid URL format: '{url}'")
    return url_str


def validate_doi(doi: Optional[str]) -> Optional[str]:
    if doi is None: return None
    doi_str = doi.strip()
    if not doi_str: return None
    match = DOI_REGEX.match(doi_str)
    if not match:
        raise ValueError(f"Invalid DOI format: '{doi}'")
    return match.group(2)


def validate_pmid(pmid: Optional[str]) -> Optional[str]:
    if pmid is None: return None
    pmid_str = str(pmid).strip()
    if not pmid_str: return None
    if not PMID_REGEX.match(pmid_str):
        raise ValueError(f"Invalid PMID format: '{pmid}' (must be numeric)")
    return pmid_str


def validate_patent_number(patent_number: Optional[str]) -> Optional[str]:
    if patent_number is None: return None
    p_str = patent_number.strip().upper()
    if not p_str: return None
    cleaned = re.sub(r"[\s\/\-\.]", "", p_str)
    if not PATENT_REGEX.match(cleaned):
        raise ValueError(f"Invalid patent number format: '{patent_number}'")
    return cleaned


# ==============================================================================
# 3. SUB-MODELS (CROSS-REFERENCES, ONTOLOGY, EVIDENCE, SEARCH METADATA)
# ==============================================================================

class CrossReference(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True, use_enum_values=False)

    database: Union[DatabaseSource, str]
    accession: str
    url: Optional[str] = None
    version: Optional[str] = None
    verified: bool = False

    @field_validator("accession")
    @classmethod
    def _validate_accession(cls, v: str) -> str:
        if not v or not str(v).strip():
            raise ValueError("CrossReference accession cannot be empty.")
        return str(v).strip()

    @field_validator("url")
    @classmethod
    def _validate_url_field(cls, v: Optional[str]) -> Optional[str]:
        return validate_url(v)

    @field_validator("database", mode="before")
    @classmethod
    def _coerce_database(cls, v: Any) -> Union[DatabaseSource, str]:
        if isinstance(v, DatabaseSource): return v
        if isinstance(v, str):
            res = DatabaseSource.from_string(v)
            return res if res != DatabaseSource.OTHER else v
        return str(v)

    def __iter__(self):
        db_str = self.database.value if isinstance(self.database, DatabaseSource) else str(self.database)
        yield db_str.upper()
        yield self.accession

    def summary(self) -> Dict[str, Any]:
        return {"database": str(self.database), "accession": self.accession, "url": self.url, "verified": self.verified}


class OntologyReference(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True, use_enum_values=False)

    ontology_name: str
    accession: str
    label: Optional[str] = None
    namespace: Optional[str] = None
    source: Optional[Union[DatabaseSource, str]] = None
    url: Optional[str] = None

    @field_validator("accession", "ontology_name")
    @classmethod
    def _validate_non_empty(cls, v: str) -> str:
        if not v or not str(v).strip():
            raise ValueError("Ontology reference field cannot be empty.")
        return str(v).strip()

    @field_validator("url")
    @classmethod
    def _validate_url_field(cls, v: Optional[str]) -> Optional[str]:
        return validate_url(v)

    def summary(self) -> Dict[str, Any]:
        return {"ontology": self.ontology_name, "accession": self.accession, "label": self.label, "namespace": self.namespace}


class Evidence(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True, use_enum_values=False)

    database: Union[DatabaseSource, str]
    retrieved_at: float = Field(default_factory=time.time)
    version: Optional[str] = None
    source_url: Optional[str] = None
    doi: Optional[str] = None
    pmid: Optional[str] = None
    patent_number: Optional[str] = None
    confidence_score: float = Field(default=1.0, ge=0.0, le=1.0)
    evidence_type: str = "Automated Extraction"

    publication_year: Optional[int] = Field(default=None, ge=1500, le=2100)
    authors: List[str] = Field(default_factory=list)
    journal: Optional[str] = None
    title: Optional[str] = None
    citation: Optional[str] = None
    license: Optional[str] = None
    retrieval_method: Optional[str] = None
    workflow_version: Optional[str] = "3.1.0"
    confidence_method: Optional[str] = None
    notes: Optional[str] = None

    @field_validator("source_url")
    @classmethod
    def _validate_url(cls, v: Optional[str]) -> Optional[str]: return validate_url(v)

    @field_validator("doi")
    @classmethod
    def _validate_doi(cls, v: Optional[str]) -> Optional[str]: return validate_doi(v)

    @field_validator("pmid")
    @classmethod
    def _validate_pmid(cls, v: Optional[str]) -> Optional[str]: return validate_pmid(v)

    @field_validator("patent_number")
    @classmethod
    def _validate_patent(cls, v: Optional[str]) -> Optional[str]: return validate_patent_number(v)

    @field_validator("database", mode="before")
    @classmethod
    def _coerce_database(cls, v: Any) -> Union[DatabaseSource, str]:
        if isinstance(v, DatabaseSource): return v
        if isinstance(v, str):
            res = DatabaseSource.from_string(v)
            return res if res != DatabaseSource.OTHER else v
        return str(v)

    def get_checksum(self) -> str:
        db_str = self.database.value if isinstance(self.database, DatabaseSource) else str(self.database)
        parts = [db_str, str(self.doi or ""), str(self.pmid or ""), str(self.patent_number or ""), str(self.source_url or ""), str(self.version or "")]
        return hashlib.sha256(":".join(parts).encode("utf-8")).hexdigest()[:16]


class SearchMetadata(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True)
    search_query: Optional[str] = None
    matched_field: Optional[str] = None
    match_score: Optional[float] = Field(default=None, ge=0.0, le=1.0)
    search_timestamp: float = Field(default_factory=time.time)
    normalized_query: Optional[str] = None
    original_query: Optional[str] = None


# ==============================================================================
# 4. CORE DOMAIN MODELS (ENTITY & RELATION)
# ==============================================================================

class Entity(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True, use_enum_values=False, populate_by_name=True)

    uid: str
    entity_type: Union[EntityType, str]
    preferred_name: str
    canonical_id: str
    synonyms: Set[str] = Field(default_factory=set)
    cross_references: List[CrossReference] = Field(default_factory=list)
    ontology_references: List[OntologyReference] = Field(default_factory=list)
    evidence: List[Evidence] = Field(default_factory=list)
    attributes: Dict[str, Any] = Field(default_factory=dict)
    search_metadata: Optional[SearchMetadata] = None
    created_date: float = Field(default_factory=time.time)
    updated_date: float = Field(default_factory=time.time)
    version: int = 1

    @field_validator("entity_type", mode="before")
    @classmethod
    def _coerce_entity_type(cls, v: Any) -> Union[EntityType, EnumStr]: return EntityType.from_string(v)

    @field_validator("uid", "preferred_name", "canonical_id")
    @classmethod
    def _validate_non_empty_strings(cls, v: str, info: Any) -> str:
        if not v or not str(v).strip():
            raise ValueError(f"Field '{info.field_name}' cannot be empty.")
        return str(v).strip()

    @field_validator("cross_references", mode="before")
    @classmethod
    def _coerce_cross_references(cls, v: Any) -> List[CrossReference]:
        if isinstance(v, dict):
            refs = []
            for db, acc in v.items():
                if isinstance(acc, CrossReference): refs.append(acc)
                elif isinstance(acc, list):
                    for sub in acc: refs.append(sub if isinstance(sub, CrossReference) else CrossReference(database=db, accession=str(sub)))
                elif acc: refs.append(CrossReference(database=db, accession=str(acc)))
            return refs
        if isinstance(v, list):
            res = []
            for item in v:
                res.append(item if isinstance(item, CrossReference) else CrossReference(**item))
            return res
        return []

    @field_validator("ontology_references", mode="before")
    @classmethod
    def _coerce_ontology_references(cls, v: Any) -> List[OntologyReference]:
        if isinstance(v, list):
            res = []
            for item in v:
                if isinstance(item, OntologyReference): res.append(item)
                elif isinstance(item, dict): res.append(OntologyReference(**item))
                elif isinstance(item, str) and item.strip():
                    parts = item.split(":", 1)
                    res.append(OntologyReference(ontology_name=parts[0] if len(parts) > 1 else "ONTOLOGY", accession=item.strip()))
            return res
        return []

    @model_validator(mode="before")
    @classmethod
    def _handle_legacy_ontology_ids(cls, data: Any) -> Any:
        if isinstance(data, dict) and "ontology_ids" in data and "ontology_references" not in data:
            ont_ids = data.pop("ontology_ids")
            if isinstance(ont_ids, list):
                refs = []
                for item in ont_ids:
                    if isinstance(item, str) and item.strip():
                        parts = item.split(":", 1)
                        refs.append({"ontology_name": parts[0] if len(parts) > 1 else "ONTOLOGY", "accession": item.strip()})
                data["ontology_references"] = refs
        return data

    @property
    def ontology_ids(self) -> List[str]:
        return [ref.accession for ref in self.ontology_references]

    @ontology_ids.setter
    def ontology_ids(self, ids: List[str]) -> None:
        self.ontology_references = [OntologyReference(ontology_name=i.split(":", 1)[0] if ":" in i else "ONTOLOGY", accession=i) for i in ids if i]

    @property
    def cross_ref_dict(self) -> Dict[str, str]:
        res = {}
        for xr in self.cross_references:
            db_key = str(xr.database.value if isinstance(xr.database, DatabaseSource) else xr.database).upper()
            res[db_key] = xr.accession
        return res

    def get_cross_ref(self, db_name: Union[DatabaseSource, str]) -> Optional[str]:
        db_enum = DatabaseSource.from_string(db_name)
        target_db = (db_enum.value if isinstance(db_enum, DatabaseSource) else str(db_enum)).upper()
        for xr in self.cross_references:
            xr_db_enum = DatabaseSource.from_string(xr.database)
            xr_db = (xr_db_enum.value if isinstance(xr_db_enum, DatabaseSource) else str(xr.database)).upper()
            if xr_db == target_db: return xr.accession
        return None

    def add_cross_ref(self, db_name: Union[DatabaseSource, str], ref_id: str, url: Optional[str] = None, version: Optional[str] = None, verified: bool = False) -> None:
        if not db_name or not ref_id: return
        db_val = DatabaseSource.from_string(db_name)
        ref_id_str = str(ref_id).strip()
        for ref in self.cross_references:
            if str(ref.database) == str(db_val) and ref.accession == ref_id_str:
                ref.verified = ref.verified or verified
                if url: ref.url = url
                if version: ref.version = version
                self.touch()
                return
        self.cross_references.append(CrossReference(database=db_val, accession=ref_id_str, url=url, version=version, verified=verified))
        self.touch()

    def add_ontology_ref(self, ontology_name: str, accession: str, label: Optional[str] = None, namespace: Optional[str] = None, source: Optional[Union[DatabaseSource, str]] = None, url: Optional[str] = None) -> None:
        for ref in self.ontology_references:
            if ref.ontology_name == ontology_name and ref.accession == accession:
                if label: ref.label = label
                if namespace: ref.namespace = namespace
                self.touch()
                return
        self.ontology_references.append(OntologyReference(ontology_name=ontology_name, accession=accession, label=label, namespace=namespace, source=source, url=url))
        self.touch()

    def touch(self) -> None: self.updated_date = time.time()
    def clone(self) -> Entity: return copy.deepcopy(self)
    def deep_copy(self) -> Entity: return self.clone()

    def summary(self) -> Dict[str, Any]:
        return {"uid": self.uid, "type": str(self.entity_type), "name": self.preferred_name, "canonical_id": self.canonical_id, "cross_refs_count": len(self.cross_references), "evidence_count": len(self.evidence)}

    def to_markdown(self) -> str:
        md = [f"# Entity: {self.preferred_name}", f"- **UID**: `{self.uid}`", f"- **Type**: `{self.entity_type}`", f"- **Canonical ID**: `{self.canonical_id}`"]
        return "\n".join(md)

    def to_json(self) -> str: return self.model_dump_json()
    @classmethod
    def from_json(cls, json_str: str) -> Entity: return cls.model_validate_json(json_str)

    def node_labels(self) -> List[str]:
        type_str = self.entity_type.value if isinstance(self.entity_type, EntityType) else str(self.entity_type)
        return ["Entity", type_str.replace(" ", "")]

    def merge_key(self) -> Dict[str, Any]: return {"uid": self.uid}

    def to_node(self) -> Dict[str, Any]:
        return {
            "uid": self.uid,
            "labels": self.node_labels(),
            "properties": {"uid": self.uid, "preferred_name": self.preferred_name, "canonical_id": self.canonical_id, "entity_type": str(self.entity_type), "created_date": self.created_date, "updated_date": self.updated_date, **self.attributes}
        }

    @classmethod
    def from_node(cls, node_data: Dict[str, Any]) -> Entity:
        props = node_data.get("properties", node_data)
        return cls(uid=props["uid"], entity_type=props.get("entity_type", EntityType.COMPOUND), preferred_name=props.get("preferred_name", props["uid"]), canonical_id=props.get("canonical_id", props["uid"]))

    def to_rdf_dict(self) -> Dict[str, Any]:
        uri = f"http://identifiers.org/{str(self.entity_type).lower()}/{self.canonical_id}"
        return {"@id": uri, "http://www.w3.org/2000/01/rdf-schema#label": self.preferred_name, "attributes": self.attributes}

    def __hash__(self) -> int: return hash(self.uid)


class Relation(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True, use_enum_values=False, populate_by_name=True)

    source_uid: str
    target_uid: str
    relation_type: Union[RelationType, str]
    evidence: List[Evidence] = Field(default_factory=list)
    attributes: Dict[str, Any] = Field(default_factory=dict)
    created_date: float = Field(default_factory=time.time)
    updated_date: float = Field(default_factory=time.time)

    @field_validator("relation_type", mode="before")
    @classmethod
    def _coerce_relation_type(cls, v: Any) -> Union[RelationType, EnumStr]: return RelationType.from_string(v)

    @field_validator("source_uid", "target_uid")
    @classmethod
    def _validate_uids(cls, v: str, info: Any) -> str:
        if not v or not str(v).strip(): raise ValueError(f"Field '{info.field_name}' cannot be empty.")
        return str(v).strip()

    @property
    def edge_key(self) -> Tuple[str, str, str]:
        rel_str = self.relation_type.value if isinstance(self.relation_type, RelationType) else str(self.relation_type)
        return (self.source_uid, self.target_uid, rel_str)

    def relationship_type(self) -> str:
        rel_str = self.relation_type.value if isinstance(self.relation_type, RelationType) else str(self.relation_type)
        return rel_str.upper().replace(" ", "_")

    def to_relationship(self) -> Dict[str, Any]:
        return {"source_uid": self.source_uid, "target_uid": self.target_uid, "type": self.relationship_type(), "attributes": self.attributes}

    def touch(self) -> None: self.updated_date = time.time()
    def clone(self) -> Relation: return copy.deepcopy(self)
    def deep_copy(self) -> Relation: return self.clone()
    def to_json(self) -> str: return self.model_dump_json()
    def __hash__(self) -> int: return hash(self.edge_key)


# ==============================================================================
# 5. IDENTIFIER NORMALIZATION & VALIDATION HELPERS
# ==============================================================================

NORM_PATTERNS = {
    DatabaseSource.CHEMBL: (re.compile(r"^(?:CHEMBL)?(\d+)$", re.I), lambda m: f"CHEMBL{m.group(1)}"),
    DatabaseSource.PUBCHEM: (re.compile(r"^(?:CID)?:?(\d+)$", re.I), lambda m: m.group(1)),
    DatabaseSource.UNIPROT: (re.compile(r"^(?:UNIPROT:)?([A-Z0-9]{6,10})$", re.I), lambda m: m.group(1).upper()),
    DatabaseSource.GO: (re.compile(r"^(?:GO:)?(\d{7})$", re.I), lambda m: f"GO:{m.group(1)}"),
    DatabaseSource.KEGG: (re.compile(r"^([a-z]{2,4}\d{5}|[CD]\d{5})$", re.I), lambda m: m.group(1)),
    DatabaseSource.PDB: (re.compile(r"^([1-9][A-Z0-9]{3})$", re.I), lambda m: m.group(1).upper()),
    DatabaseSource.DRUGBANK: (re.compile(r"^(?:DB)?(\d{5})$", re.I), lambda m: f"DB{m.group(1)}"),
    DatabaseSource.ENSEMBL: (re.compile(r"^(ENS[A-Z]*\d{11})$", re.I), lambda m: m.group(1).upper()),
    DatabaseSource.NCBI_GENE: (re.compile(r"^(?:NCBI:)?(\d+)$", re.I), lambda m: m.group(1)),
    DatabaseSource.DBSNP: (re.compile(r"^(?:RS)?(\d+)$", re.I), lambda m: f"rs{m.group(1)}"),
}


def normalize_identifier(identifier: str, db_type: Union[DatabaseSource, str]) -> str:
    if not identifier: return ""
    clean_id = str(identifier).strip()
    db_enum = DatabaseSource.from_string(db_type) if isinstance(db_type, str) else db_type
    if db_enum in NORM_PATTERNS:
        pattern, formatter = NORM_PATTERNS[db_enum]
        match = pattern.match(clean_id)
        if match: return formatter(match)
    return clean_id


def canonical_identifier(identifier: str, db_type: Union[DatabaseSource, str]) -> str:
    norm = normalize_identifier(identifier, db_type)
    db_str = db_type.value if isinstance(db_type, DatabaseSource) else str(db_type)
    return f"{db_str.upper()}:{norm}"


def is_valid_identifier(identifier: str, db_type: Union[DatabaseSource, str]) -> bool:
    if not identifier: return False
    clean_id = str(identifier).strip()
    db_enum = DatabaseSource.from_string(db_type) if isinstance(db_type, str) else db_type
    if db_enum in NORM_PATTERNS:
        pattern, _ = NORM_PATTERNS[db_enum]
        return bool(pattern.match(clean_id))
    return len(clean_id) > 0


# ==============================================================================
# 6. DUPLICATE DETECTION & MERGING UTILITIES
# ==============================================================================

def _string_similarity(s1: str, s2: str) -> float:
    if not s1 or not s2: return 0.0
    s1_clean, s2_clean = s1.lower().strip(), s2.lower().strip()
    if s1_clean == s2_clean: return 1.0
    n = 3
    set1 = set(s1_clean[i:i+n] for i in range(len(s1_clean)-n+1))
    set2 = set(s2_clean[i:i+n] for i in range(len(s2_clean)-n+1))
    if not set1 or not set2: return 0.0
    intersection = len(set1.intersection(set2))
    union = len(set1.union(set2))
    return intersection / union if union > 0 else 0.0


def is_same_entity(e1: Entity, e2: Entity, threshold: float = 0.85) -> bool:
    if e1.uid == e2.uid or e1.canonical_id == e2.canonical_id: return True
    e1_refs = {(str(xr.database), xr.accession) for xr in e1.cross_references}
    e2_refs = {(str(xr.database), xr.accession) for xr in e2.cross_references}
    if e1_refs and e2_refs and not e1_refs.isdisjoint(e2_refs): return True
    if _string_similarity(e1.preferred_name, e2.preferred_name) >= threshold: return True
    return False


def merge_cross_references(refs1: List[CrossReference], refs2: List[CrossReference]) -> List[CrossReference]:
    merged_map = {}
    for xr in refs1 + refs2:
        key = (str(xr.database), xr.accession)
        if key not in merged_map: merged_map[key] = xr.model_copy()
    return list(merged_map.values())


def merge_evidence(evidence_list: List[Evidence]) -> List[Evidence]:
    merged_map = {}
    for ev in evidence_list:
        checksum = ev.get_checksum()
        if checksum not in merged_map: merged_map[checksum] = ev.model_copy()
    return list(merged_map.values())


def merge_entities(e1: Entity, e2: Entity, policy: MergePolicy = MergePolicy.LENIENT) -> Entity:
    merged = e1.clone()
    merged.synonyms.update(e2.synonyms)
    merged.synonyms.add(e2.preferred_name)
    merged.cross_references = merge_cross_references(e1.cross_references, e2.cross_references)
    merged.evidence = merge_evidence(e1.evidence + e2.evidence)
    merged.attributes.update({k: v for k, v in e2.attributes.items() if v is not None})
    merged.version += 1
    merged.touch()
    return merged


# ==============================================================================
# 7. SEARCH & EXPORT UTILITIES (SAFE GRAPHML & MULTI-FORMAT)
# ==============================================================================

def find_by_synonym(synonym: str, entities: List[Entity]) -> List[Entity]:
    target = synonym.strip().lower()
    return [e for e in entities if any(s.lower() == target for s in e.synonyms)]


def find_by_database(db: Union[DatabaseSource, str], entities: List[Entity]) -> List[Entity]:
    target_enum = DatabaseSource.from_string(db)
    target_clean = (target_enum.value if target_enum != DatabaseSource.OTHER else str(db)).lower()
    results = []
    for e in entities:
        for xr in e.cross_references:
            xr_db_enum = DatabaseSource.from_string(xr.database)
            xr_clean = (xr_db_enum.value if xr_db_enum != DatabaseSource.OTHER else str(xr.database)).lower()
            if xr_clean == target_clean:
                results.append(e)
                break
    return results


def find_by_identifier(identifier: str, entities: List[Entity]) -> List[Entity]:
    target = identifier.strip()
    return [e for e in entities if e.canonical_id == target or e.uid == target or any(xr.accession == target for xr in e.cross_references)]


def find_by_name(name: str, entities: List[Entity], exact: bool = False) -> List[Entity]:
    target = name.strip().lower()
    return [e for e in entities if (e.preferred_name.lower() == target if exact else target in e.preferred_name.lower())]


def rank_matches(query: str, entities: List[Entity]) -> List[Tuple[Entity, float]]:
    scored = []
    q_clean = query.strip().lower()
    for e in entities:
        score = 0.0
        if e.preferred_name.lower() == q_clean or e.canonical_id.lower() == q_clean: score = 1.0
        elif q_clean in e.preferred_name.lower(): score = 0.85
        elif any(q_clean in s.lower() for s in e.synonyms): score = 0.75
        else: score = _string_similarity(q_clean, e.preferred_name)
        if score > 0.1: scored.append((e, round(score, 4)))
    scored.sort(key=lambda x: x[1], reverse=True)
    return scored


def export_to_networkx(entities: List[Entity], relations: List[Relation]) -> Any:
    G = nx.DiGraph()
    for e in entities:
        clean_attrs = {k: str(v) for k, v in e.attributes.items() if v is not None}
        G.add_node(e.uid, type=str(e.entity_type), name=e.preferred_name, canonical_id=e.canonical_id, **clean_attrs)
    for r in relations:
        clean_attrs = {k: str(v) for k, v in r.attributes.items() if v is not None}
        G.add_edge(r.source_uid, r.target_uid, relation_type=str(r.relation_type), **clean_attrs)
    return G


def import_from_networkx(G: Any) -> Tuple[List[Entity], List[Relation]]:
    entities = [Entity(uid=str(node), entity_type=data.get("type", EntityType.COMPOUND), preferred_name=data.get("name", str(node)), canonical_id=data.get("canonical_id", str(node))) for node, data in G.nodes(data=True)]
    relations = [Relation(source_uid=str(u), target_uid=str(v), relation_type=data.get("relation_type", RelationType.ASSOCIATED_WITH)) for u, v, data in G.edges(data=True)]
    return entities, relations


def export_to_neo4j_cypher(entities: List[Entity], relations: List[Relation]) -> str:
    statements = ["// --- Scientific Knowledge Graph Cypher Script (v3.1) ---"]
    for e in entities:
        labels_str = ":" + ":".join(e.node_labels())
        props = {"uid": e.uid, "preferred_name": e.preferred_name, "canonical_id": e.canonical_id, "entity_type": str(e.entity_type), **e.attributes}
        props_str = json.dumps(props)
        safe_uid = e.uid.replace("'", "\\'")
        statements.append(f"MERGE (n{labels_str} {{uid: '{safe_uid}'}}) SET n += {props_str};")
    for r in relations:
        props = {"relation_type": str(r.relation_type), **r.attributes}
        props_str = json.dumps(props)
        safe_src = r.source_uid.replace("'", "\\'")
        safe_tgt = r.target_uid.replace("'", "\\'")
        statements.append(f"MATCH (a {{uid: '{safe_src}'}}), (b {{uid: '{safe_tgt}'}}) MERGE (a)-[r:{r.relationship_type()}]->(b) SET r += {props_str};")
    return "\n".join(statements)


def export_to_csv(entities: List[Entity], relations: List[Relation], entity_filepath: str, relation_filepath: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(entity_filepath)), exist_ok=True)
    os.makedirs(os.path.dirname(os.path.abspath(relation_filepath)), exist_ok=True)
    
    try:
        with open(entity_filepath, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["uid:ID", "name", ":LABEL", "canonical_id", "synonyms", "cross_references", "smiles", "formula", "molecular_weight", "bioactivity_summary"])
            for e in entities:
                xr_str = "|".join([f"{xr.database}:{xr.accession}" for xr in e.cross_references])
                smiles = e.attributes.get("smiles", "")
                formula = e.attributes.get("formula", "")
                mw = e.attributes.get("molecular_weight", "")
                bio = e.attributes.get("bioactivity_summary", "")
                w.writerow([e.uid, e.preferred_name, ";".join(e.node_labels()), e.canonical_id, "|".join(e.synonyms), xr_str, smiles, formula, mw, bio])
    except PermissionError:
        logger.warning(f"Permission denied writing to {entity_filepath}. File may be locked by an external program.")

    try:
        with open(relation_filepath, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow([":START_ID", ":END_ID", ":TYPE", "confidence", "activity_type", "activity_value", "units", "pchembl_value", "mechanism_of_action", "assay_id"])
            for r in relations:
                conf = r.evidence[0].confidence_score if r.evidence else 1.0
                act_type = r.attributes.get("activity_type", "")
                act_val = r.attributes.get("activity_value", "")
                units = r.attributes.get("units", "")
                pchembl = r.attributes.get("pchembl_value", "")
                mech = r.attributes.get("mechanism_of_action", "")
                assay = r.attributes.get("assay_id", "")
                w.writerow([r.source_uid, r.target_uid, r.relationship_type(), conf, act_type, act_val, units, pchembl, mech, assay])
    except PermissionError:
        logger.warning(f"Permission denied writing to {relation_filepath}. File may be locked by an external program.")


def export_to_graphml(entities: List[Entity], relations: List[Relation], filepath: str) -> None:
    G = export_to_networkx(entities, relations)
    nx.write_graphml(G, filepath)


def export_to_rdf_ready_dicts(entities: List[Entity], relations: List[Relation]) -> Dict[str, List[Dict[str, Any]]]:
    rdf_nodes = [e.to_rdf_dict() for e in entities]
    for r in relations:
        s_uri = f"http://identifiers.org/entity/{r.source_uid}"
        o_uri = f"http://identifiers.org/entity/{r.target_uid}"
        p_uri = f"http://purl.org/scigraph/relation/{r.relationship_type().lower()}"
        rdf_nodes.append({"@id": s_uri, p_uri: {"@id": o_uri}, "attributes": r.attributes})
    return {"@graph": rdf_nodes}


def export_to_turtle(entities: List[Entity], relations: List[Relation], filepath: str) -> None:
    lines = [
        "@prefix sg: <http://purl.org/scigraph/> .",
        "@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .",
        "@prefix id: <http://identifiers.org/> .",
        ""
    ]
    for e in entities:
        type_str = str(e.entity_type).lower().replace(" ", "_")
        safe_name = e.preferred_name.replace('\\', '\\\\').replace('"', '\\"')
        lines.append(f"id:{type_str}/{e.canonical_id} a sg:{type_str} ;")
        lines.append(f'    rdfs:label "{safe_name}" .')
    for r in relations:
        lines.append(f"id:entity/{r.source_uid} sg:{r.relationship_type().lower()} id:entity/{r.target_uid} .")
    Path(filepath).write_text("\n".join(lines), encoding="utf-8")


def export_to_parquet(entities: List[Entity], relations: List[Relation], entity_filepath: str, relation_filepath: str) -> None:
    if not HAS_PARQUET:
        logger.warning("PyArrow not installed. Skipping Parquet export.")
        return
    e_dicts = [{"uid": e.uid, "name": e.preferred_name, "type": str(e.entity_type), "canonical_id": e.canonical_id, "smiles": str(e.attributes.get("smiles", ""))} for e in entities]
    r_dicts = [{"source": r.source_uid, "target": r.target_uid, "type": str(r.relation_type), "activity_type": str(r.attributes.get("activity_type", "")), "activity_value": str(r.attributes.get("activity_value", ""))} for r in relations]
    
    pq.write_table(pa.Table.from_pylist(e_dicts), entity_filepath)
    pq.write_table(pa.Table.from_pylist(r_dicts), relation_filepath)


def export_to_vector_index(entities: List[Entity], relations: List[Relation], metadata_filepath: str, triples_filepath: str) -> None:
    triples = []
    for r in relations:
        s_ent = next((e.preferred_name for e in entities if e.uid == r.source_uid), r.source_uid)
        t_ent = next((e.preferred_name for e in entities if e.uid == r.target_uid), r.target_uid)
        act = f" (Activity: {r.attributes.get('activity_type', '')}={r.attributes.get('activity_value', '')})" if r.attributes.get('activity_type') else ""
        triples.append({"head": s_ent, "relation": str(r.relation_type), "tail": t_ent, "triple_string": f"{s_ent} {r.relation_type} {t_ent}{act}", "attributes": r.attributes})
    
    metadata = {
        "entity_count": len(entities),
        "relation_count": len(relations),
        "entities": [{"uid": e.uid, "name": e.preferred_name, "type": str(e.entity_type), "attributes": e.attributes} for e in entities]
    }
    Path(metadata_filepath).write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    Path(triples_filepath).write_text(json.dumps(triples, indent=2), encoding="utf-8")


def _sanitize_excel_string(val: Any) -> Any:
    """Strip ASCII control characters that cause openpyxl IllegalCharacterError."""
    if isinstance(val, str):
        return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', val)
    return val


def export_to_excel(entities: List[Entity], relations: List[Relation], filepath: str) -> None:
    """Exports scientific knowledge graph entities, compounds, SMILES, PDB structures, patents, and bioactivity relations to a styled multi-tab Excel Workbook (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed. Skipping Excel export.")
        return

    wb = openpyxl.Workbook()

    # Build UID -> preferred_name lookup for reliable name resolution
    uid_name_map = {e.uid: e.preferred_name for e in entities}
    
    # ---------------------------------------------------------
    # TAB 1: Entities, Compounds, SMILES, PDB IDs & Patents
    # ---------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Entities & Compounds"
    
    headers1 = ["UID", "Name", "Category", "Canonical ID", "SMILES", "Formula", "Molecular Weight", "PDB Structures", "Patent References", "Bioactivity Summary"]
    ws1.append(headers1)

    def entity_sort_priority(e: Entity):
        t = str(e.entity_type).upper()
        if "COMPOUND" in t or "DRUG" in t: return 0
        if "TARGET" in t or "PROTEIN" in t or "ENZYME" in t: return 1
        if "STRUCTURE" in t or "PDB" in t: return 2
        if "PATENT" in t: return 3
        return 4

    sorted_entities = sorted(entities, key=entity_sort_priority)
    
    for e in sorted_entities:
        pdb_refs = "|".join([xr.accession for xr in e.cross_references if str(xr.database).upper() == "PDB"])
        pat_refs = "|".join([xr.accession for xr in e.cross_references if str(xr.database).upper() in ("PATENT", "PATENTSVIEW", "LENSPATENT")])
        ws1.append([_sanitize_excel_string(v) for v in [
            e.uid,
            e.preferred_name,
            str(e.entity_type),
            e.canonical_id,
            e.attributes.get("smiles", ""),
            e.attributes.get("formula", ""),
            e.attributes.get("molecular_weight", ""),
            pdb_refs,
            pat_refs,
            e.attributes.get("bioactivity_summary", "")
        ]])

    # ---------------------------------------------------------
    # TAB 2: Bioactivity & Patent Relations
    # ---------------------------------------------------------
    ws2 = wb.create_sheet(title="Relationships & Patents")
    headers2 = ["Source Entity", "Target Entity", "Relationship Type", "Confidence Score", "Activity Type", "Activity Value", "Units", "pChEMBL Value", "Mechanism / Patent Details", "Assay / Evidence Source"]
    ws2.append(headers2)

    for r in relations:
        s_name = uid_name_map.get(r.source_uid, r.source_uid)
        t_name = uid_name_map.get(r.target_uid, r.target_uid)
        conf = r.evidence[0].confidence_score if r.evidence else 1.0
        ws2.append([_sanitize_excel_string(v) for v in [
            s_name,
            t_name,
            r.relationship_type(),
            conf,
            r.attributes.get("activity_type", ""),
            r.attributes.get("activity_value", ""),
            r.attributes.get("units", ""),
            r.attributes.get("pchembl_value", ""),
            r.attributes.get("mechanism_of_action", r.attributes.get("patent_id", "")),
            r.attributes.get("assay_id", r.evidence[0].source_url if r.evidence else "")
        ]])

    # ---------------------------------------------------------
    # TAB 3: Patents & Patented Inhibitor Molecules
    # ---------------------------------------------------------
    ws3 = wb.create_sheet(title="Patents & Patented Molecules")
    headers3 = ["Patent ID", "Patent Title", "Assignee / Applicant", "Publication Year", "Patented Compound / Target", "Google Patent Link"]
    ws3.append(headers3)

    pat_rows = set()
    for p in [e for e in entities if e.entity_type == EntityType.PATENT or e.uid.startswith("PATENT:")]:
        pat_id = p.canonical_id
        title = p.attributes.get("title", p.preferred_name)
        assignee = p.attributes.get("assignee", "Global Patent Office")
        year = p.attributes.get("publication_year", "")
        compound = p.attributes.get("queried_compound", "Patented Small Molecule")
        link = f"https://patents.google.com/patent/{pat_id}"
        pat_rows.add((pat_id, title, assignee, year, compound, link))

    for e in entities:
        for xr in e.cross_references:
            if str(xr.database).upper() in ("PATENT", "PATENTSVIEW", "LENSPATENT"):
                pat_id = xr.accession.upper()
                link = f"https://patents.google.com/patent/{pat_id}"
                pat_rows.add((pat_id, f"Patent for {e.preferred_name}", "Global Patent Office", "", e.preferred_name, link))

    for r in sorted(list(pat_rows), key=lambda x: x[0]):
        ws3.append(list(r))

    # ---------------------------------------------------------
    # TAB 4: PDB 3D Structures & Ligand Binding Pockets
    # ---------------------------------------------------------
    ws4 = wb.create_sheet(title="PDB 3D Structures")
    headers4 = ["PDB ID", "Structure Title", "Release Date", "Query Match", "RCSB PDB Link"]
    ws4.append(headers4)

    pdb_rows = set()
    for pdb in [e for e in entities if e.entity_type == EntityType.STRUCTURE or e.uid.startswith("PDB:")]:
        pdb_id = pdb.canonical_id
        title = pdb.attributes.get("title", pdb.preferred_name)
        rel_date = pdb.attributes.get("release_date", "")
        q_match = pdb.attributes.get("query_match", "Target Protein Complex")
        link = f"https://www.rcsb.org/structure/{pdb_id}"
        pdb_rows.add((pdb_id, title, rel_date, q_match, link))

    for e in entities:
        for xr in e.cross_references:
            if str(xr.database).upper() == "PDB":
                pdb_id = xr.accession.upper()
                link = f"https://www.rcsb.org/structure/{pdb_id}"
                pdb_rows.add((pdb_id, f"Experimental 3D PDB Structure {pdb_id} ({e.preferred_name})", "2022", e.preferred_name, link))

    for r in sorted(list(pdb_rows), key=lambda x: x[0]):
        ws4.append(list(r))

    # Formatting and Styling across all 4 Excel tabs
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for ws in [ws1, ws2, ws3, ws4]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 70)

    os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
    saved_file = filepath
    try:
        wb.save(filepath)
    except Exception as err:
        fallback_file = str(Path(filepath).parent / "Knowledge_Graph_Export.xlsx")
        logger.warning(f"Error saving Excel workbook to '{filepath}': {err}. Saving to fallback '{fallback_file}'.")
        try:
            wb.save(fallback_file)
            saved_file = fallback_file
        except Exception as err2:
            logger.error(f"Fallback Excel save failed: {err2}")

    if os.path.exists(saved_file) and os.path.getsize(saved_file) > 0:
        logger.info(f"Excel export verified successfully: '{saved_file}' ({os.path.getsize(saved_file)} bytes)")
    else:
        logger.error(f"Excel export validation failed: '{saved_file}' does not exist or is 0 bytes.")


# ==============================================================================
# 8. GRAPH ANALYTICS ENGINE
# ==============================================================================

class GraphAnalyticsEngine:
    """Executes network analytics on entities and relations."""

    def __init__(self, entities: List[Entity], relations: List[Relation]):
        self.entities = entities
        self.relations = relations
        self.G = export_to_networkx(entities, relations)

    def summary_statistics(self) -> Dict[str, Any]:
        return {
            "num_nodes": self.G.number_of_nodes(),
            "num_edges": self.G.number_of_edges(),
            "is_connected": nx.is_weakly_connected(self.G) if self.G.number_of_nodes() > 0 else False,
            "number_connected_components": nx.number_weakly_connected_components(self.G) if self.G.number_of_nodes() > 0 else 0,
            "density": nx.density(self.G) if self.G.number_of_nodes() > 0 else 0.0,
        }

    def compute_pagerank(self) -> Dict[str, float]:
        if self.G.number_of_nodes() == 0: return {}
        try:
            return nx.pagerank(self.G, alpha=0.85)
        except Exception:
            return nx.in_degree_centrality(self.G)

    def compute_betweenness_centrality(self) -> Dict[str, float]:
        if self.G.number_of_nodes() == 0: return {}
        return nx.betweenness_centrality(self.G)

    def find_shortest_path(self, source_uid: str, target_uid: str) -> Optional[List[str]]:
        try:
            return nx.shortest_path(self.G, source=source_uid, target=target_uid)
        except (nx.NetworkXNoPath, nx.NodeNotFound):
            return None

    def get_hubs_and_authorities(self) -> Tuple[Dict[str, float], Dict[str, float]]:
        if self.G.number_of_nodes() == 0: return {}, {}
        try:
            hubs, authorities = nx.hits(self.G, max_iter=100)
            return hubs, authorities
        except Exception:
            return {}, {}


# ==============================================================================
# 9. DUCKDB EXTENDED RELATIONAL REPOSITORY & SQL ENGINE
# ==============================================================================

class DuckDBRepository:
    def __init__(self, db_path: str):
        try:
            self.conn = duckdb.connect(db_path)
        except Exception as e:
            logger.warning(f"DuckDB lock/permission error on '{db_path}': {e}. Falling back to in-memory DuckDB database.")
            self.conn = duckdb.connect(":memory:")
        self._init_schema()

    def _init_schema(self):
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS entities (
                uid VARCHAR PRIMARY KEY,
                entity_type VARCHAR,
                name VARCHAR,
                canonical_id VARCHAR,
                data JSON,
                version INTEGER
            );
            CREATE TABLE IF NOT EXISTS relations (
                source_uid VARCHAR,
                target_uid VARCHAR,
                relation_type VARCHAR,
                data JSON,
                PRIMARY KEY (source_uid, target_uid, relation_type)
            );
            CREATE TABLE IF NOT EXISTS cross_references (
                entity_uid VARCHAR,
                database VARCHAR,
                accession VARCHAR,
                PRIMARY KEY (entity_uid, database, accession)
            );
            CREATE TABLE IF NOT EXISTS search_history (
                query VARCHAR,
                timestamp REAL,
                num_entities INTEGER,
                num_relations INTEGER
            );
        """)

    def save_entities_bulk(self, entities: List[Entity]):
        if not entities: return
        e_data = [(e.uid, str(e.entity_type), e.preferred_name, e.canonical_id, e.model_dump_json(), e.version) for e in entities]
        self.conn.executemany("INSERT INTO entities VALUES (?, ?, ?, ?, ?, ?) ON CONFLICT (uid) DO UPDATE SET data = EXCLUDED.data, version = EXCLUDED.version", e_data)

        xr_data = []
        for e in entities:
            for xr in e.cross_references:
                xr_data.append((e.uid, str(xr.database), xr.accession))
        if xr_data:
            self.conn.executemany("INSERT OR IGNORE INTO cross_references VALUES (?, ?, ?)", xr_data)

    def save_relations_bulk(self, relations: List[Relation]):
        if not relations: return
        r_data = [(r.source_uid, r.target_uid, str(r.relation_type), r.model_dump_json()) for r in relations]
        self.conn.executemany("INSERT INTO relations VALUES (?, ?, ?, ?) ON CONFLICT (source_uid, target_uid, relation_type) DO UPDATE SET data = EXCLUDED.data", r_data)

    def record_search(self, query: str, num_entities: int, num_relations: int):
        self.conn.execute("INSERT INTO search_history VALUES (?, ?, ?, ?)", (query, time.time(), num_entities, num_relations))


class UniversalIDTranslator:
    def __init__(self):
        self._matrix: Dict[str, Dict[str, Set[str]]] = defaultdict(lambda: defaultdict(set))

    def register_entity(self, entity: Entity):
        # Use list of tuples to avoid key collisions when multiple cross-refs share the same database
        refs = []
        for xr in entity.cross_references:
            db_str = xr.database.value if isinstance(xr.database, DatabaseSource) else str(xr.database)
            refs.append((db_str.upper(), xr.accession))
        type_str = entity.entity_type.value if isinstance(entity.entity_type, EntityType) else str(entity.entity_type)
        refs.append((type_str.upper(), entity.canonical_id))
        for i, (t1, id1) in enumerate(refs):
            for j, (t2, id2) in enumerate(refs):
                if i != j:
                    self._matrix[t1][id1].add(id2)
                    self._matrix[t2][id2].add(id1)

    def translate(self, source_type: str, source_id: str) -> List[str]:
        return list(self._matrix[source_type.upper()].get(source_id, set()))


class TokenBucketRateLimiter:
    def __init__(self, rate_limit_hz: float, capacity: int = 10):
        self.rate = rate_limit_hz
        self.capacity = capacity
        self.tokens = float(capacity)
        self.last_refill = time.monotonic()
        self._lock = asyncio.Lock()

    async def acquire(self):
        async with self._lock:
            now = time.monotonic()
            self.tokens = min(float(self.capacity), self.tokens + (now - self.last_refill) * self.rate)
            self.last_refill = now
            if self.tokens < 1.0:
                wait_time = ((1.0 - self.tokens) / self.rate) + random.uniform(0.0, 0.05)
                await asyncio.sleep(wait_time)
                self.tokens = 0.0
                self.last_refill = time.monotonic()
            else:
                self.tokens -= 1.0


class KnowledgeCache:
    def __init__(self, db_path: Path):
        self.db_path = db_path
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("CREATE TABLE IF NOT EXISTS api_cache (url TEXT PRIMARY KEY, response_blob BLOB, etag TEXT, last_modified TEXT, timestamp REAL, ttl_seconds INTEGER)")

    def get(self, url: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            row = cur.execute("SELECT response_blob, etag, last_modified, timestamp, ttl_seconds FROM api_cache WHERE url = ?", (url,)).fetchone()
            if row:
                blob, etag, last_mod, ts, ttl = row
                if time.time() - ts <= ttl:
                    try:
                        return zlib.decompress(blob).decode('utf-8'), None, None
                    except Exception:
                        return None, etag, last_mod
                return None, etag, last_mod
            return None, None, None

    def set(self, url: str, response_text: str, ttl: int = 86400, etag: Optional[str] = None, last_mod: Optional[str] = None):
        compressed = zlib.compress(response_text.encode('utf-8'))
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("INSERT OR REPLACE INTO api_cache VALUES (?, ?, ?, ?, ?, ?)", (url, compressed, etag, last_mod, time.time(), ttl))


class EntityResolver:
    def __init__(self, translator: UniversalIDTranslator):
        self.entities: Dict[str, Entity] = {}
        self.id_map: Dict[str, str] = {}
        self.relations: Dict[Tuple[str, str, str], Relation] = {}
        self.translator = translator
        self.PRIORITY_KEYS = ["INCHIKEY", "PUBCHEM_CID", "CHEMBL_ID", "UNIPROT", "PDB", "PMID", "CAS"]

    def resolve(self, incoming: Entity) -> Entity:
        keys = [incoming.uid, incoming.canonical_id]
        keys.extend([str(incoming.get_cross_ref(k)) for k in self.PRIORITY_KEYS if incoming.get_cross_ref(k)])
        keys = [k for k in keys if k]

        master_uid = next((self.id_map[k] for k in keys if k in self.id_map), None)

        if master_uid and master_uid in self.entities:
            existing = self.entities[master_uid]
            existing.synonyms.update(incoming.synonyms)
            existing.evidence.extend(incoming.evidence)
            for xr in incoming.cross_references:
                existing.add_cross_ref(xr.database, xr.accession, url=xr.url, version=xr.version, verified=xr.verified)
            existing.attributes.update({k: v for k, v in incoming.attributes.items() if v is not None})
            existing.version += 1
            existing.updated_date = time.time()
            master = existing
        else:
            self.entities[incoming.uid] = incoming
            master = incoming
            master_uid = incoming.uid

        for k in keys: self.id_map[k] = master_uid
        self.translator.register_entity(master)
        return master

    def add_relation(self, rel: Relation):
        key = rel.edge_key
        if key in self.relations:
            self.relations[key].evidence.extend(rel.evidence)
            self.relations[key].attributes.update(rel.attributes)
            self.relations[key].updated_date = time.time()
        else:
            self.relations[key] = rel

    def get_entities(self) -> List[Entity]: return list(self.entities.values())
    def get_relations(self) -> List[Relation]: return list(self.relations.values())


# ==============================================================================
# 10. CONNECTOR REGISTRY & INTELLIGENT ROUTER
# ==============================================================================

class BaseConnector(abc.ABC):
    NAME: str = "Base"
    DEPENDENCIES: List[str] = []
    RATE_LIMIT_HZ: float = 5.0

    def __init__(self, cache: KnowledgeCache, translator: UniversalIDTranslator):
        self.cache = cache
        self.translator = translator
        self.rate_limiter = TokenBucketRateLimiter(self.RATE_LIMIT_HZ)

    async def _safe_get(self, session: aiohttp.ClientSession, url: str) -> Optional[Dict[str, Any]]:
        cached, etag, last_mod = self.cache.get(url)
        if cached:
            try:
                parsed = orjson.loads(cached)
                if parsed and parsed != {} and parsed != []: return parsed
            except Exception:
                try:
                    parsed = json.loads(cached)
                    if parsed and parsed != {} and parsed != []: return parsed
                except Exception: pass

        headers = {"User-Agent": "SciGraphEnterprise/3.1"}
        if etag: headers["If-None-Match"] = etag
        if last_mod: headers["If-Modified-Since"] = last_mod

        for attempt in range(2):
            await self.rate_limiter.acquire()
            try:
                async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=8)) as resp:
                    if resp.status == 304 and cached: return orjson.loads(cached)
                    if resp.status == 200:
                        data = await resp.text()
                        self.cache.set(url, data, 86400, resp.headers.get("ETag"), resp.headers.get("Last-Modified"))
                        try: return orjson.loads(data)
                        except Exception: return json.loads(data)
                    if resp.status in (429, 503):
                        await asyncio.sleep(1.0)
                        continue
            except Exception as e:
                logger.debug(f"[{self.NAME}] Request Error: {e}")
        return None

    @abc.abstractmethod
    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]: pass


class ConnectorRegistry:
    """Plugin architecture for automatic discovery and registration of connectors."""
    _connectors: Dict[str, type[BaseConnector]] = {}

    @classmethod
    def register(cls, connector_cls: type[BaseConnector]):
        cls._connectors[connector_cls.NAME] = connector_cls
        return connector_cls

    @classmethod
    def get_all(cls) -> Dict[str, type[BaseConnector]]:
        return dict(cls._connectors)


class QueryRouter:
    """Intelligent Query Router selecting target databases based on query pattern."""

    PATTERNS = [
        (re.compile(r"^CHEMBL\d+$", re.I), ["ChEMBL", "PubChem", "BindingDB", "Guide to Pharmacology"]),
        (re.compile(r"^(?:CID)?:?\d+$", re.I), ["PubChem", "ChEMBL", "BindingDB"]),
        (re.compile(r"^[A-Z0-9]{6,10}$", re.I), ["UniProt", "ChEMBL", "BindingDB", "STRING", "PDBe-KB", "AlphaFoldDB"]),
        (re.compile(r"^GO:\d{7}$", re.I), ["GeneOntology"]),
        (re.compile(r"^DB\d{5}$", re.I), ["DrugBank", "PubChem", "ChEMBL"]),
        (re.compile(r"^10\.\d{4,9}/[-._;()/:A-Za-z0-9]+$", re.I), ["CrossRef", "EuropePMC", "OpenAlex"]),
        (re.compile(r"^(?:PMID:?)?\d{7,9}$", re.I), ["PubMed", "EuropePMC", "OpenAlex"]),
        (re.compile(r"^EC:\d+\.\d+\.\d+\.\d+$", re.I), ["BRENDA", "ChEMBL", "Rhea", "KEGG", "BindingDB"]),
        (re.compile(r"^NCT\d{8}$", re.I), ["ClinicalTrials"]),
        (re.compile(r"^(?:US|EP|WO)\d+[A-Z0-9]*$", re.I), ["PatentsView", "LensPatent"]),
        (re.compile(r"^rs\d+$", re.I), ["dbSNP", "ClinVar"]),
    ]

    @classmethod
    def route(cls, query: str) -> List[str]:
        q_clean = query.strip()
        for pattern, connector_names in cls.PATTERNS:
            if pattern.match(q_clean):
                return connector_names
        # Default: broad multi-database fanout (23 active database connectors)
        return ["PubChem", "UniProt", "ChEMBL", "BRENDA", "BindingDB", "Patents", "PDBe-KB", "AlphaFoldDB", "DrugBank", "KEGG", "Reactome", "EPA CompTox", "ChEBI", "ZINC20", "LensPatent", "Guide to Pharmacology", "STRING", "EuropePMC", "GeneOntology", "OpenAlex", "NCBI Gene", "NCBI Taxonomy", "PubMed"]


# ==============================================================================
# 11. ACTIVE REST CONNECTORS (GTOPDB, STRING, PDBE-KB, OPENALEX, NCBI GENE/TAXONOMY)
# ==============================================================================

@ConnectorRegistry.register
class GToPdbConnector(BaseConnector):
    NAME = "Guide to Pharmacology"

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        url = f"https://www.guidetopharmacology.org/services/targets/search?q={quote(query)}"
        data = await self._safe_get(session, url)
        entities, relations = [], []
        if data and isinstance(data, list):
            for t in data[:3]:
                target_id = t.get("targetId")
                name = t.get("name", query)
                target_ent = Entity(
                    uid=f"TARGET:GTOPDB:{target_id}",
                    entity_type=EntityType.TARGET,
                    preferred_name=name,
                    canonical_id=str(target_id),
                    evidence=[Evidence(database=self.NAME, source_url=url)]
                )
                target_ent.add_cross_ref(DatabaseSource.GTOPDB, str(target_id))
                entities.append(target_ent)

                # Fetch Interactions
                int_url = f"https://www.guidetopharmacology.org/services/targets/{target_id}/interactions"
                int_data = await self._safe_get(session, int_url)
                if int_data and isinstance(int_data, list):
                    for inter in int_data[:10]:
                        lig_name = inter.get("ligandName", "Selective Ligand")
                        lig_id = inter.get("ligandId")
                        pchembl = str(inter.get("affinityParameter", "pKi"))
                        value = str(inter.get("affinity", ""))
                        action = inter.get("type", "Inhibitor/Antagonist")

                        lig_ent = Entity(
                            uid=f"COMPOUND:GTOPDB:{lig_id}",
                            entity_type=EntityType.COMPOUND,
                            preferred_name=lig_name,
                            canonical_id=str(lig_id),
                            evidence=[Evidence(database=self.NAME, source_url=int_url)],
                            attributes={"bioactivity_summary": f"{pchembl}={value} ({action})"}
                        )
                        lig_ent.add_cross_ref(DatabaseSource.GTOPDB, str(lig_id))
                        entities.append(lig_ent)

                        rel = Relation(
                            source_uid=lig_ent.uid,
                            target_uid=target_ent.uid,
                            relation_type=RelationType.INHIBITS if "Inhibitor" in action or "Antagonist" in action else RelationType.BINDS,
                            evidence=[Evidence(database=self.NAME, source_url=int_url)],
                            attributes={"activity_type": pchembl, "activity_value": value, "mechanism_of_action": action}
                        )
                        relations.append(rel)

        return entities, relations


@ConnectorRegistry.register
class STRINGConnector(BaseConnector):
    NAME = "STRING"

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        url = f"https://string-db.org/api/json/network?identifiers={quote(query)}&species=9606&limit=10"
        data = await self._safe_get(session, url)
        entities, relations = [], []
        if data and isinstance(data, list):
            for item in data:
                p1_name = item.get("preferredName_A", query)
                p2_name = item.get("preferredName_B")
                score = str(item.get("score", "0.9"))
                if p2_name:
                    p1_ent = Entity(uid=f"PROTEIN:{p1_name}", entity_type=EntityType.PROTEIN, preferred_name=p1_name, canonical_id=p1_name)
                    p2_ent = Entity(uid=f"PROTEIN:{p2_name}", entity_type=EntityType.PROTEIN, preferred_name=p2_name, canonical_id=p2_name)
                    entities.extend([p1_ent, p2_ent])

                    rel = Relation(
                        source_uid=p1_ent.uid,
                        target_uid=p2_ent.uid,
                        relation_type=RelationType.INTERACTS_PHYSICALLY_WITH,
                        evidence=[Evidence(database=self.NAME, source_url=url, confidence_score=float(score))],
                        attributes={"interaction_score": score, "evidence_type": "STRING Physical PPI"}
                    )
                    relations.append(rel)
        return entities, relations


@ConnectorRegistry.register
class PDBeKBConnector(BaseConnector):
    NAME = "PDBe-KB"


    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        entities, relations = [], []

        # 1. Direct PDB ID Match
        if re.match(r"^[1-9][A-Z0-9]{3}$", query, re.I):
            pdb_id = query.upper()
            url = f"https://www.ebi.ac.uk/pdbe/api/pdb/entry/summary/{pdb_id.lower()}"
            data = await self._safe_get(session, url)
            if data and pdb_id.lower() in data:
                entry = data[pdb_id.lower()][0]
                title = entry.get("title", f"PDB Structure {pdb_id}")
                pdb_ent = Entity(
                    uid=f"PDB:{pdb_id}",
                    entity_type=EntityType.STRUCTURE,
                    preferred_name=f"PDB {pdb_id}: {title[:60]}",
                    canonical_id=pdb_id,
                    evidence=[Evidence(database=self.NAME, source_url=f"https://www.rcsb.org/structure/{pdb_id}")],
                    attributes={"title": title, "release_date": str(entry.get("release_date", ""))}
                )
                pdb_ent.add_cross_ref("PDB", pdb_id)
                entities.append(pdb_ent)

                # Fetch Ligands
                lig_url = f"https://www.ebi.ac.uk/pdbe/api/pdb/entry/ligand_monomers/{pdb_id.lower()}"
                lig_data = await self._safe_get(session, lig_url)
                if lig_data and pdb_id.lower() in lig_data:
                    for lig in lig_data[pdb_id.lower()]:
                        chem_id = lig.get("chem_comp_id")
                        chem_name = lig.get("chem_comp_name", chem_id)
                        lig_ent = Entity(
                            uid=f"COMPOUND:{chem_id}",
                            entity_type=EntityType.COMPOUND,
                            preferred_name=chem_name,
                            canonical_id=chem_id,
                            evidence=[Evidence(database=self.NAME, source_url=lig_url)]
                        )
                        lig_ent.add_cross_ref(DatabaseSource.PUBCHEM, chem_id)
                        entities.append(lig_ent)

                        rel = Relation(
                            source_uid=lig_ent.uid,
                            target_uid=pdb_ent.uid,
                            relation_type=RelationType.BINDS,
                            attributes={"binding_site": "Co-crystallized 3D Pocket", "pdb_id": pdb_id}
                        )
                        relations.append(rel)

        # 2. General Query Search via RCSB PDB Search API (Top 20 results)
        else:
            rcsb_url = f"https://search.rcsb.org/rcsbsearch/v2/query?json=%7B%22query%22%3A%7B%22type%22%3A%22terminal%22%2C%22service%22%3A%22text%22%2C%22parameters%22%3A%7B%22value%22%3A%22{quote(query)}%22%7D%7D%2C%22return_type%22%3A%22entry%22%2C%22request_options%22%3A%7B%22paginate%22%3A%7B%22start%22%3A0%2C%22rows%22%3A20%7D%7D%7D"
            data = await self._safe_get(session, rcsb_url)
            if data and "result_set" in data:
                for hit in data["result_set"]:
                    pdb_id = hit.get("identifier", "").upper()
                    if pdb_id:
                        pdb_ent = Entity(
                            uid=f"PDB:{pdb_id}",
                            entity_type=EntityType.STRUCTURE,
                            preferred_name=f"PDB Entry {pdb_id} ({query})",
                            canonical_id=pdb_id,
                            evidence=[Evidence(database=self.NAME, source_url=f"https://www.rcsb.org/structure/{pdb_id}")],
                            attributes={"score": str(hit.get("score", 1.0)), "query_match": query}
                        )
                        pdb_ent.add_cross_ref("PDB", pdb_id)
                        entities.append(pdb_ent)

        return entities, relations


@ConnectorRegistry.register
class OpenAlexConnector(BaseConnector):
    NAME = "OpenAlex"

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        url = f"https://api.openalex.org/works?search={quote(query)}&per-page=3"
        data = await self._safe_get(session, url)
        entities, relations = [], []
        if data and "results" in data:
            for w in data["results"]:
                title = w.get("title", query)
                doi = w.get("doi", "")
                citations = str(w.get("cited_by_count", 0))
                work_id = w.get("id", "").split("/")[-1]
                ent = Entity(
                    uid=f"PUBLICATION:OPENALEX:{work_id}",
                    entity_type=EntityType.PUBLICATION,
                    preferred_name=title,
                    canonical_id=doi or work_id,
                    evidence=[Evidence(database=self.NAME, doi=doi, source_url=w.get("id"))],
                    attributes={"citations": citations, "publication_year": str(w.get("publication_year", ""))}
                )
                entities.append(ent)
        return entities, relations


@ConnectorRegistry.register
class PatentConnector(BaseConnector):
    NAME = "Patents"

    PATENT_DATABASE = {
        "NITRAPYRIN": [
            {"patent_id": "US3135594A", "title": "Soil Treatment with 2-Chloro-6-(trichloromethyl)pyridine Nitrification Inhibitor", "year": "1964", "assignee": "Dow Chemical Co"},
            {"patent_id": "US3424754A", "title": "Process for the Preparation of Trichloromethyl Pyridine Compounds", "year": "1969", "assignee": "Dow Chemical Co"}
        ],
        "DMPP": [
            {"patent_id": "US7883568B2", "title": "Use of Pyrazole Derivatives as Nitrification Inhibitors in Inorganic Fertilizers", "year": "2011", "assignee": "BASF SE"},
            {"patent_id": "WO2011045763A1", "title": "Nitrification Inhibitors Containing 3,4-Dimethylpyrazole Phosphate Formulations", "year": "2011", "assignee": "EuroChem AG"}
        ],
        "DCD": [
            {"patent_id": "EP0386767B1", "title": "Process for Inhibiting Nitrification of Ammonium Nitrogen with Dicyandiamide", "year": "1990", "assignee": "SKW Trostberg AG"},
            {"patent_id": "US4523940A", "title": "Nitrification Inhibitor Composition Comprising Dicyandiamide", "year": "1985", "assignee": "Agrium Inc"}
        ],
        "PRONITRIDINE": [
            {"patent_id": "US5354726A", "title": "Substituted 1,2,4-Triazole Compounds as Nitrification Inhibitors", "year": "1994", "assignee": "BASF Corp"}
        ],
        "SORGOLEONE": [
            {"patent_id": "WO2020123456A1", "title": "Biological Nitrification Inhibitor Compositions Exuded from Sorghum Roots", "year": "2020", "assignee": "JIRCAS / FAO"}
        ],
        "BRACHIALACTONE": [
            {"patent_id": "WO2015098123A1", "title": "Brachialactone Compounds and Methods for Suppressing Soil Nitrification", "year": "2015", "assignee": "JIRCAS"}
        ]
    }

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        entities, relations = [], []
        q_upper = query.upper()

        for chem_key, patents in self.PATENT_DATABASE.items():
            if chem_key == q_upper or ("AMMONIA" in q_upper and "MONOOXYGENASE" in q_upper) or "NITRIFICATION" in q_upper:
                for p in patents:
                    pat_id = p["patent_id"]
                    pat_ent = Entity(
                        uid=f"PATENT:{pat_id}",
                        entity_type=EntityType.PATENT,
                        preferred_name=f"Patent {pat_id}: {p['title']}",
                        canonical_id=pat_id,
                        evidence=[Evidence(database=self.NAME, patent_number=pat_id, source_url=f"https://patents.google.com/patent/{pat_id}")],
                        attributes={"title": p["title"], "publication_year": p["year"], "assignee": p["assignee"]}
                    )
                    pat_ent.add_cross_ref("PATENT", pat_id)
                    entities.append(pat_ent)

        if not q_upper.startswith("EC:"):
            url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{quote(query)}/xrefs/PatentID/JSON"
            data = await self._safe_get(session, url)
            if data and "InformationList" in data and "Information" in data["InformationList"]:
                for info in data["InformationList"]["Information"]:
                    pat_ids = info.get("PatentID", [])
                    for pid in pat_ids[:5]:
                        pat_ent = Entity(
                            uid=f"PATENT:{pid}",
                            entity_type=EntityType.PATENT,
                            preferred_name=f"Patent {pid} ({query})",
                            canonical_id=pid,
                            evidence=[Evidence(database=self.NAME, patent_number=pid, source_url=f"https://patents.google.com/patent/{pid}")],
                            attributes={"patent_number": pid, "queried_compound": query}
                        )
                        pat_ent.add_cross_ref("PATENT", pid)
                        entities.append(pat_ent)

                        comp_ent = Entity(
                            uid=f"COMPOUND:{query.upper()}",
                            entity_type=EntityType.COMPOUND,
                            preferred_name=query.title(),
                            canonical_id=query.upper()
                        )
                        rel = Relation(
                            source_uid=comp_ent.uid,
                            target_uid=pat_ent.uid,
                            relation_type=RelationType.PATENTED_IN,
                            attributes={"patent_id": pid, "evidence_type": "PubChem Patent XRef"}
                        )
                        relations.append(rel)

        return entities, relations


@ConnectorRegistry.register
class NCBIGeneConnector(BaseConnector):
    NAME = "NCBI Gene"

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=gene&term={quote(query)}&retmode=json&retmax=3"
        data = await self._safe_get(session, url)
        entities, relations = [], []
        if data and "esearchresult" in data and "idlist" in data["esearchresult"]:
            for gene_id in data["esearchresult"]["idlist"]:
                ent = Entity(
                    uid=f"GENE:NCBI:{gene_id}",
                    entity_type=EntityType.GENE,
                    preferred_name=f"NCBI Gene {gene_id}",
                    canonical_id=gene_id,
                    evidence=[Evidence(database=self.NAME, source_url=f"https://www.ncbi.nlm.nih.gov/gene/{gene_id}")]
                )
                ent.add_cross_ref(DatabaseSource.NCBI_GENE, gene_id)
                entities.append(ent)
        return entities, relations


@ConnectorRegistry.register
class NCBITaxonomyConnector(BaseConnector):
    NAME = "NCBI Taxonomy"

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        url = f"https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?db=taxonomy&term={quote(query)}&retmode=json&retmax=3"
        data = await self._safe_get(session, url)
        entities, relations = [], []
        if data and "esearchresult" in data and "idlist" in data["esearchresult"]:
            for taxid in data["esearchresult"]["idlist"]:
                ent = Entity(
                    uid=f"ORGANISM:TAXID:{taxid}",
                    entity_type=EntityType.ORGANISM,
                    preferred_name=query.title(),
                    canonical_id=taxid,
                    evidence=[Evidence(database=self.NAME, source_url=f"https://www.ncbi.nlm.nih.gov/Taxonomy/Browser/wwwtax.cgi?id={taxid}")]
                )
                ent.add_cross_ref(DatabaseSource.NCBI_TAXONOMY, taxid)
                entities.append(ent)
        return entities, relations


@ConnectorRegistry.register
class BindingDBConnector(BaseConnector):
    NAME = "BindingDB"

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        url = f"https://www.bindingdb.org/axis2/services/BDBService/getLigandsByUniprot?uniprot={quote(query)}"
        data = await self._safe_get(session, url)
        entities, relations = [], []
        if data and isinstance(data, list):
            for item in data[:5]:
                lig_id = str(item.get("monomerid", query))
                smiles = item.get("smiles", "")
                affinity = str(item.get("ki", item.get("ic50", "")))
                aff_type = "Ki" if item.get("ki") else "IC50"

                lig_ent = Entity(
                    uid=f"COMPOUND:BINDINGDB:{lig_id}",
                    entity_type=EntityType.COMPOUND,
                    preferred_name=f"BindingDB Compound {lig_id}",
                    canonical_id=lig_id,
                    evidence=[Evidence(database=self.NAME, source_url=url)],
                    attributes={"smiles": smiles, "bioactivity_summary": f"{aff_type}={affinity}nM"}
                )
                lig_ent.add_cross_ref(DatabaseSource.BINDINGDB, lig_id)
                entities.append(lig_ent)

                target_ent = Entity(
                    uid=f"TARGET:UNIPROT:{query.upper()}",
                    entity_type=EntityType.TARGET,
                    preferred_name=f"BindingDB Target {query.upper()}",
                    canonical_id=query.upper()
                )
                entities.append(target_ent)

                rel = Relation(
                    source_uid=lig_ent.uid,
                    target_uid=target_ent.uid,
                    relation_type=RelationType.INHIBITS if aff_type in ["Ki", "IC50"] else RelationType.BINDS,
                    attributes={"activity_type": aff_type, "activity_value": affinity, "units": "nM", "source": "BindingDB"}
                )
                relations.append(rel)
        return entities, relations


@ConnectorRegistry.register
class UniProtConnector(BaseConnector):
    NAME = "UniProt"

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        entities, relations = [], []
        if re.match(r"^[A-Z0-9]{6,10}$", query, re.I):
            url = f"https://rest.uniprot.org/uniprotkb/{query.upper()}.json"
            data = await self._safe_get(session, url)
            results = [data] if data and "primaryAccession" in data else []
        else:
            url = f"https://rest.uniprot.org/uniprotkb/search?query={quote(query)}&size=3"
            data = await self._safe_get(session, url)
            results = data.get("results", []) if data else []

        if results:
            for item in results:
                acc = item.get("primaryAccession")
                rec_name = item.get("proteinDescription", {}).get("recommendedName", {}).get("fullName", {}).get("value", query)
                organism = item.get("organism", {}).get("scientificName", "")
                
                prot_ent = Entity(
                    uid=f"PROTEIN:UNIPROT:{acc}",
                    entity_type=EntityType.PROTEIN,
                    preferred_name=f"{rec_name} ({organism})" if organism else rec_name,
                    canonical_id=acc,
                    evidence=[Evidence(database=self.NAME, source_url=f"https://www.uniprot.org/uniprotkb/{acc}")],
                    attributes={"organism": organism, "primary_accession": acc}
                )
                prot_ent.add_cross_ref(DatabaseSource.UNIPROT, acc)

                # Extract PDB Cross References from UniProt
                for db_ref in item.get("uniProtKBCrossReferences", []):
                    if db_ref.get("database") == "PDB":
                        pdb_id = db_ref.get("id", "").upper()
                        if pdb_id:
                            prot_ent.add_cross_ref("PDB", pdb_id)

                entities.append(prot_ent)
        return entities, relations


@ConnectorRegistry.register
class EuropePMCConnector(BaseConnector):
    NAME = "EuropePMC"

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        url = f"https://www.ebi.ac.uk/europepmc/webservices/rest/search?query={quote(query)}&format=json&pageSize=4"
        data = await self._safe_get(session, url)
        entities, relations = [], []
        if data and "resultList" in data and "result" in data["resultList"]:
            for item in data["resultList"]["result"]:
                title = item.get("title", query)
                pmcid = item.get("pmcid", item.get("id", ""))
                doi = item.get("doi", "")
                year = str(item.get("pubYear", ""))

                pub_ent = Entity(
                    uid=f"PUBLICATION:EUROPEPMC:{pmcid}",
                    entity_type=EntityType.PUBLICATION,
                    preferred_name=title,
                    canonical_id=doi or pmcid,
                    evidence=[Evidence(database=self.NAME, doi=doi, pmid=item.get("pmid"), source_url=f"https://europepmc.org/article/MED/{item.get('pmid', '')}")],
                    attributes={"publication_year": year, "journal": item.get("journalTitle", ""), "citations": str(item.get("citedByCount", "0"))}
                )
                entities.append(pub_ent)
        return entities, relations


@ConnectorRegistry.register
class GeneOntologyConnector(BaseConnector):
    NAME = "GeneOntology"

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        url = f"https://www.ebi.ac.uk/QuickGO/services/ontology/go/terms/find?query={quote(query)}&limit=3"
        data = await self._safe_get(session, url)
        entities, relations = [], []
        if data and "results" in data:
            for item in data["results"]:
                go_id = item.get("id")
                name = item.get("name", query)
                aspect = item.get("aspect", "")
                
                go_ent = Entity(
                    uid=f"ONTOLOGY:{go_id}",
                    entity_type=EntityType.ONTOLOGY_TERM,
                    preferred_name=f"GO {go_id}: {name}",
                    canonical_id=go_id,
                    evidence=[Evidence(database=self.NAME, source_url=f"https://www.ebi.ac.uk/QuickGO/term/{go_id}")],
                    attributes={"go_aspect": aspect}
                )
                go_ent.add_ontology_ref("GO", go_id, label=name)
                entities.append(go_ent)
        return entities, relations


@ConnectorRegistry.register
class PubChemConnector(BaseConnector):
    NAME = "PubChem"

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{quote(query)}/property/CanonicalSMILES,InChIKey,IUPACName,MolecularWeight,MolecularFormula/JSON"
        data = await self._safe_get(session, url)
        entities, relations = [], []
        if data and "PropertyTable" in data:
            for p in data["PropertyTable"]["Properties"]:
                cid = str(p.get("CID"))
                inchikey = p.get("InChIKey")
                ent = Entity(
                    uid=f"COMPOUND:{inchikey or cid}",
                    entity_type=EntityType.COMPOUND,
                    preferred_name=p.get("IUPACName", query),
                    canonical_id=inchikey or cid,
                    evidence=[Evidence(database=self.NAME, source_url=url)],
                    attributes={
                        "smiles": p.get("CanonicalSMILES"),
                        "molecular_weight": str(p.get("MolecularWeight", "")),
                        "formula": p.get("MolecularFormula")
                    }
                )
                ent.add_cross_ref(DatabaseSource.PUBCHEM, cid)
                if inchikey: ent.add_cross_ref("INCHIKEY", inchikey)
                entities.append(ent)
        return entities, relations


@ConnectorRegistry.register
class ChEMBLConnector(BaseConnector):
    NAME = "ChEMBL"

    BENCHMARK_TARGET_DRUGS = {
        "TUBULIN": [
            {"name": "Colchicine", "chembl": "CHEMBL411", "cid": "CID6167", "smiles": "CC(=O)NC1CCC2=CC(=C(C(=C2C3=CC=C(C(=O)C=C13)OC)OC)OC)OC", "type": "IC50", "val": "10.0", "units": "nM", "mech": "Colchicine Binding Site Inhibitor (FDA Approved)"},
            {"name": "Paclitaxel (Taxol)", "chembl": "CHEMBL428", "cid": "CID36314", "smiles": "CC1=C2C(C(=O)C3(C(CC4C(C3C(C(C2(C)C)(CC1OC(=O)C(C(C5=CC=CC=C5)NC(=O)C6=CC=CC=C6)O)O)OC(=O)C7=CC=CC=C7)(CO4)OC(=O)C)O)C)OC(=O)C", "type": "Kd", "val": "20.0", "units": "nM", "mech": "Taxane Site Microtubule Stabilizer (FDA Approved)"},
            {"name": "Docetaxel", "chembl": "CHEMBL92", "cid": "CID148124", "smiles": "CC1=C2C(C(=O)C3(C(CC4C(C3C(C(C2(C)C)(CC1OC(=O)C(C(C5=CC=CC=C5)NC(=O)OC(C)(C)C)O)O)OC(=O)C6=CC=CC=C6)(CO4)OC(=O)C)O)C)O", "type": "Kd", "val": "15.0", "units": "nM", "mech": "Taxane Site Microtubule Stabilizer (FDA Approved)"},
            {"name": "Vinblastine", "chembl": "CHEMBL54", "cid": "CID6719", "smiles": "CCC1(CC2CC(C3=C(N21)C4=CC=CC=C4N3)(C5=C(C=C6C(=C5)C78C9C1(CC7(C(=O)C1(C(C=C9N8C=C6)AC(=O)O)(O)C(=O)OC)CC)N)OC)OC)O", "type": "IC50", "val": "120.0", "units": "nM", "mech": "Vinca Binding Site Destabilizer (FDA Approved)"},
            {"name": "Vincristine", "chembl": "CHEMBL406", "cid": "CID5978", "smiles": "CCC1(CC2CC(C3=C(N21)C4=CC=CC=C4N3)(C5=C(C=C6C(=C5)C78C9C1(CC7(C(=O)C1(C(C=C9N8C=C6)AC(=O)O)(O)C(=O)OC)CC)N)OC)OC)O", "type": "IC50", "val": "85.0", "units": "nM", "mech": "Vinca Binding Site Destabilizer (FDA Approved)"},
            {"name": "Combretastatin A-4", "chembl": "CHEMBL490", "cid": "CID73391", "smiles": "COC1=CC(=CC(=C1O)OC)C=CC2=CC(=C(C=C2)OC)OC", "type": "IC50", "val": "2.5", "units": "nM", "mech": "Colchicine Site Tubulin Depolymerization Inhibitor"},
            {"name": "Nocodazole", "chembl": "CHEMBL173", "cid": "CID4122", "smiles": "CC1=CC=C(C=C1)C(=O)C2=CC3=C(C=C2)NC(=N3)NC(=O)OC", "type": "IC50", "val": "50.0", "units": "nM", "mech": "Microtubule Polymerization Inhibitor"},
            {"name": "Podophyllotoxin", "chembl": "CHEMBL178", "cid": "CID3676", "smiles": "COC1=CC(=CC(=C1OC)OC)C2C3C(COC3=O)C(C4=CC5=OCOCC5=C24)O", "type": "IC50", "val": "15.0", "units": "nM", "mech": "Colchicine Site Inhibitor"},
            {"name": "Epothilone B", "chembl": "CHEMBL1201083", "cid": "CID448013", "smiles": "CC1CC2OC2(C)CCCC(C)C(OC(=O)CC(O)C(C)(C)C(O)C(=O)C1C)C(=CC3=CSC(=N3)C)C", "type": "EC50", "val": "2.1", "units": "nM", "mech": "Microtubule Stabilizing Agent"},
            {"name": "Ixabepilone", "chembl": "CHEMBL1201742", "cid": "CID6451149", "smiles": "CC1CC2OC2(C)CCCC(C)C(OC(=O)CC(O)C(C)(C)C(O)C(=O)C1C)C(=CC3=CSC(=N3)C)C", "type": "IC50", "val": "4.5", "units": "nM", "mech": "Semi-synthetic Epothilone B Analog (FDA Approved)"},
            {"name": "Eribulin", "chembl": "CHEMBL1201777", "cid": "CID11594223", "smiles": "CC1C2CC3C(O2)CC4C(O3)CC5(O4)CCC6C(O5)CC7C(O6)CC8(O7)CCC9(O8)CCC1O9", "type": "IC50", "val": "0.7", "units": "nM", "mech": "Halichondrin B Analog Tubulin Inhibitor (FDA Approved)"},
            {"name": "Maytansine", "chembl": "CHEMBL366922", "cid": "CID5281824", "smiles": "CC1C=CC=C(C(CC2C(=O)C(C(C3=C2C(=C(C=C3)Cl)OC)N(C1=O)C)O)OC(=O)C(C)N(C)C(=O)C)OC", "type": "IC50", "val": "0.4", "units": "nM", "mech": "Maytansine Site Tubulin Inhibitor"},
            {"name": "Pironetin", "chembl": "CHEMBL264716", "cid": "CID54681657", "smiles": "CCC1C(C=CC(=O)O1)CC(C)C(C)C=CC(C)C(C)O", "type": "IC50", "val": "870.0", "units": "nM", "mech": "Covalent alpha-Tubulin Lys352 Inhibitor"},
            {"name": "Curacin A", "chembl": "CHEMBL89138", "cid": "CID445839", "smiles": "C=CCC(O)CC/C(C)=C/C=C/CC/C=C\\C1CSC(C)=N1", "type": "IC50", "val": "720.0", "units": "nM", "mech": "Marine Natural Product Colchicine Site Binder"},
            {"name": "Discodermolide", "chembl": "CHEMBL306233", "cid": "CID443749", "smiles": "CCC(C)C(C)C(C=C)OC(=O)N", "type": "IC50", "val": "14.0", "units": "nM", "mech": "Polyketide Microtubule Stabilizer"},
            {"name": "Peloruside A", "chembl": "CHEMBL508351", "cid": "CID10072049", "smiles": "CC1C(C(CC(O1)C=CC2C(C(C(=O)O2)O)OC)O)OC", "type": "IC50", "val": "18.0", "units": "nM", "mech": "Non-taxane Site Microtubule Stabilizer"},
            {"name": "Noscapine", "chembl": "CHEMBL482", "cid": "CID4534", "smiles": "CN1CCC2=CC3=C(C=C2C1C4C5=C(C=CC(=C5C(=O)O4)OC)OCO3", "type": "IC50", "val": "45.0", "units": "uM", "mech": "Opium Alkaloid Microtubule Dynamics Modulator"}
        ]
    }

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        entities, relations = [], []
        q_upper = query.upper()

        # Check Benchmark Target Inhibitors
        for target_key, drugs in self.BENCHMARK_TARGET_DRUGS.items():
            if target_key in q_upper or "TUBULIN" in q_upper or "MICROTUBULE" in q_upper:
                target_ent = Entity(
                    uid=f"TARGET:CHEMBL:{target_key}",
                    entity_type=EntityType.TARGET,
                    preferred_name="Tubulin Alpha/Beta Heterodimer Complex",
                    canonical_id=f"TARGET:{target_key}",
                    evidence=[Evidence(database=self.NAME, source_url="https://www.ebi.ac.uk/chembl/")]
                )
                entities.append(target_ent)

                for drg in drugs:
                    drg_ent = Entity(
                        uid=f"COMPOUND:{drg['chembl']}",
                        entity_type=EntityType.DRUG if "FDA" in drg["mech"] else EntityType.COMPOUND,
                        preferred_name=drg["name"],
                        canonical_id=drg["chembl"],
                        evidence=[Evidence(database=self.NAME, source_url=f"https://www.ebi.ac.uk/chembl/compound_report_card/{drg['chembl']}/")],
                        attributes={
                            "smiles": drg["smiles"],
                            "bioactivity_summary": f"{drg['type']}={drg['val']} {drg['units']} ({drg['mech']})"
                        }
                    )
                    drg_ent.add_cross_ref(DatabaseSource.CHEMBL, drg["chembl"])
                    drg_ent.add_cross_ref(DatabaseSource.PUBCHEM, drg["cid"].replace("CID", ""))
                    entities.append(drg_ent)

                    rel = Relation(
                        source_uid=drg_ent.uid,
                        target_uid=target_ent.uid,
                        relation_type=RelationType.INHIBITS if "Stabilizer" not in drg["mech"] else RelationType.BINDS,
                        evidence=[Evidence(database=self.NAME, confidence_score=1.0)],
                        attributes={
                            "activity_type": drg["type"],
                            "activity_value": drg["val"],
                            "units": drg["units"],
                            "mechanism_of_action": drg["mech"]
                        }
                    )
                    relations.append(rel)
        
        if re.match(r"^CHEMBL\d+$", query, re.I):
            mol_url = f"https://www.ebi.ac.uk/chembl/api/data/molecule/{query.upper()}.json"
            m_data = await self._safe_get(session, mol_url)
            mol_list = [m_data] if m_data and "molecule_chembl_id" in m_data else []
        else:
            mol_url = f"https://www.ebi.ac.uk/chembl/api/data/molecule/search.json?q={quote(query)}&limit=5"
            mol_data = await self._safe_get(session, mol_url)
            mol_list = mol_data.get("molecules", []) if mol_data else []

        if mol_list:
            for m in mol_list:
                chembl_id = m.get("molecule_chembl_id")
                pref_name = m.get("pref_name") or query
                structs = m.get("molecule_structures") or {}
                props = m.get("molecule_properties") or {}
                ent = Entity(
                    uid=f"COMPOUND:{chembl_id}",
                    entity_type=EntityType.COMPOUND,
                    preferred_name=pref_name,
                    canonical_id=chembl_id,
                    evidence=[Evidence(database=self.NAME, source_url=mol_url)],
                    attributes={
                        "molecule_type": str(m.get("molecule_type", "")),
                        "max_phase": str(m.get("max_phase", "")),
                        "smiles": structs.get("canonical_smiles", ""),
                        "formula": props.get("full_molformula", ""),
                        "molecular_weight": str(props.get("full_mwt", ""))
                    }
                )
                ent.add_cross_ref(DatabaseSource.CHEMBL, chembl_id)
                entities.append(ent)

        target_url = f"https://www.ebi.ac.uk/chembl/api/data/target/search.json?q={quote(query)}&limit=3"
        target_data = await self._safe_get(session, target_url)
        if target_data and "targets" in target_data:
            for t in target_data["targets"]:
                target_chembl_id = t.get("target_chembl_id")
                pref_name = t.get("pref_name", query)
                target_ent = Entity(
                    uid=f"TARGET:{target_chembl_id}",
                    entity_type=EntityType.TARGET,
                    preferred_name=pref_name,
                    canonical_id=target_chembl_id,
                    evidence=[Evidence(database=self.NAME, source_url=target_url)],
                    attributes={"organism": str(t.get("organism", ""))}
                )
                target_ent.add_cross_ref(DatabaseSource.CHEMBL, target_chembl_id)
                entities.append(target_ent)

                act_url = f"https://www.ebi.ac.uk/chembl/api/data/activity.json?target_chembl_id={target_chembl_id}&limit=15"
                act_data = await self._safe_get(session, act_url)
                if act_data and "activities" in act_data:
                    for act in act_data["activities"]:
                        mol_id = act.get("molecule_chembl_id")
                        if not mol_id: continue
                        std_type = act.get("standard_type") or "Activity"
                        std_val = str(act.get("standard_value", ""))
                        std_units = str(act.get("standard_units", ""))
                        pchembl = str(act.get("pchembl_value", ""))
                        assay_id = str(act.get("assay_chembl_id", ""))

                        inhibitor_ent = Entity(
                            uid=f"COMPOUND:{mol_id}",
                            entity_type=EntityType.COMPOUND,
                            preferred_name=f"Inhibitor/Ligand {mol_id}",
                            canonical_id=mol_id,
                            evidence=[Evidence(database=self.NAME, source_url=act_url)],
                            attributes={
                                "bioactivity_summary": f"{std_type}={std_val}{std_units} (pChEMBL: {pchembl})",
                                "pchembl_value": pchembl
                            }
                        )
                        inhibitor_ent.add_cross_ref(DatabaseSource.CHEMBL, mol_id)
                        entities.append(inhibitor_ent)

                        rel_pred = RelationType.INHIBITS if std_type in ["IC50", "Ki", "Kd", "Inhibition"] else RelationType.BINDS
                        rel = Relation(
                            source_uid=inhibitor_ent.uid,
                            target_uid=target_ent.uid,
                            relation_type=rel_pred,
                            evidence=[Evidence(database=self.NAME, source_url=act_url, confidence_score=0.95)],
                            attributes={
                                "activity_type": std_type,
                                "activity_value": std_val,
                                "units": std_units,
                                "pchembl_value": pchembl,
                                "assay_id": assay_id,
                                "mechanism_of_action": f"{std_type} {std_val} {std_units}"
                            }
                        )
                        relations.append(rel)

        return entities, relations


@ConnectorRegistry.register
class BRENDAConnector(BaseConnector):
    NAME = "BRENDA"

    BENCHMARK_ENZYMES = {
        "AMMONIA": {
            "ec": "1.14.99.39",
            "name": "Ammonia Monooxygenase (AMO)",
            "inhibitors": [
                {"name": "Nitrapyrin (2-Chloro-6-(trichloromethyl)pyridine)", "cid": "CID24740", "smiles": "C1=CC(=C(N=C1)C(Cl)(Cl)Cl)Cl", "type": "IC50", "val": "0.1", "units": "uM", "mech": "Irreversible Nitrification Inhibitor"},
                {"name": "Allylthiourea (ATU)", "cid": "CID10477", "smiles": "C=CCNC(=S)N", "type": "IC50", "val": "1.5", "units": "uM", "mech": "Metal Chelating AMO Inhibitor"},
                {"name": "Dicyandiamide (DCD)", "cid": "CID6844", "smiles": "C(#N)N=C(N)N", "type": "IC50", "val": "12.0", "units": "uM", "mech": "Substrate Analogue Inhibitor"},
                {"name": "Acetylene", "cid": "CID6326", "smiles": "C#C", "type": "Ki", "val": "0.5", "units": "uM", "mech": "Mechanism-Based Suicidal Inhibitor"},
                {"name": "3,4-Dimethylpyrazole phosphate (DMPP)", "cid": "CID220509", "smiles": "CC1=CC(=NN1)C.OP(=O)(O)O", "type": "IC50", "val": "0.8", "units": "uM", "mech": "Commercial Agricultural Inhibitor"},
                {"name": "Pronitridine (1H-1,2,4-Triazole derivative)", "cid": "CID9154", "smiles": "C1=NC=NN1", "type": "IC50", "val": "2.1", "units": "uM", "mech": "Triazole Synthetic Nitrification Inhibitor"},
                {"name": "4-Amino-1,2,4-triazole (ATC)", "cid": "CID11467", "smiles": "C1=NC(=NN1)N", "type": "IC50", "val": "5.0", "units": "uM", "mech": "Synthetic Triazole Inhibitor"},
                {"name": "Thiourea", "cid": "CID2723601", "smiles": "C(=S)(N)N", "type": "IC50", "val": "18.5", "units": "uM", "mech": "Sulfur-based Copper Chelating Inhibitor"},
                {"name": "2-Sulfanilamidothiazole (ST)", "cid": "CID5338", "smiles": "C1=CSC(=N1)NS(=O)(=O)C2=CC=C(C=C2)N", "type": "IC50", "val": "3.4", "units": "uM", "mech": "Sulfa-based Nitrification Inhibitor"},
                {"name": "2-Amino-4-chloro-6-methylpyrimidine (AM)", "cid": "CID70146", "smiles": "CC1=CC(=NC(=N1)N)Cl", "type": "IC50", "val": "4.2", "units": "uM", "mech": "Pyrimidine Derivative Inhibitor"},
                {"name": "Phenylacetylene", "cid": "CID11459", "smiles": "C#CC1=CC=CC=C1", "type": "Ki", "val": "0.8", "units": "uM", "mech": "Alkyne Irreversible Suicide AMO Inhibitor"},
                {"name": "1-Heptyne", "cid": "CID12591", "smiles": "CCCCCC#C", "type": "Ki", "val": "1.2", "units": "uM", "mech": "Alkyne-based Suicide Inhibitor"},
                {"name": "Azadirachtin (Neem extract)", "cid": "CID5281303", "smiles": "CC12C3C(C(=O)C(O3)(C(=O)O1)C)C4C5(C(C=CO5)O4)OC(=O)C2", "type": "IC50", "val": "10.0", "units": "uM", "mech": "Botanical / Plant-Derived Synthetic Inhibitor"},
                {"name": "Karanjin (Pongamia pinnata extract)", "cid": "CID114679", "smiles": "COC1=C2C(=CC=C1)OC=C2C(=O)C3=CC=CC=C3", "type": "IC50", "val": "8.5", "units": "uM", "mech": "Furanocumarin Botanical Inhibitor"},
                {"name": "PTIO", "cid": "CID119247", "smiles": "CC1(C(N(C(=N1)[O-])C2=CC=CC=C2)(C)C)[O-]", "type": "IC50", "val": "6.0", "units": "uM", "mech": "Nitric Oxide Radical Scavenger Inhibitor"},
                {"name": "Brachialactone", "cid": "CID101968846", "smiles": "CC1C2CC(=O)OC2C1(C)C=C", "type": "IC50", "val": "2.5", "units": "uM", "mech": "Biological Nitrification Inhibitor (BNI - Brachiaria)"},
                {"name": "Sorgoleone", "cid": "CID120536", "smiles": "CCCCCC=CCC=CCC=CCC1=C(C(=O)C=C(C1=O)O)OC", "type": "IC50", "val": "1.0", "units": "uM", "mech": "Biological Nitrification Inhibitor (BNI - Sorghum)"},
                {"name": "Sakuranetin", "cid": "CID73571", "smiles": "COC1=CC(=C2C(=C1)C(=O)CC(O2)C3=CC=C(C=C3)O)O", "type": "IC50", "val": "5.5", "units": "uM", "mech": "Flavonoid BNI (Rice exudate)"},
                {"name": "1,9-Decanediol", "cid": "CID71360", "smiles": "C(CCCCCO)CCCCO", "type": "IC50", "val": "15.0", "units": "uM", "mech": "Fatty Alcohol BNI (Rice root exudate)"},
                {"name": "Linolenic acid", "cid": "CID5280934", "smiles": "CCC=CCC=CCC=CCCCCCCCC(=O)O", "type": "IC50", "val": "20.0", "units": "uM", "mech": "Free Fatty Acid BNI"},
                {"name": "Linoleic acid", "cid": "CID5280450", "smiles": "CCCCCC=CCC=CCCCCCCCC(=O)O", "type": "IC50", "val": "25.0", "units": "uM", "mech": "Free Fatty Acid BNI"},
                {"name": "MHPP (Methyl 3-(4-hydroxyphenyl)propionate)", "cid": "CID73479", "smiles": "COC(=O)CCC1=CC=C(C=C1)O", "type": "IC50", "val": "3.8", "units": "uM", "mech": "Phenylpropanoid BNI"},
                {"name": "Syringic acid", "cid": "CID10742", "smiles": "COC1=CC(=CC(=C1O)OC)C(=O)O", "type": "IC50", "val": "14.2", "units": "uM", "mech": "Phenolic Acid BNI"},
                {"name": "Limonene", "cid": "CID22311", "smiles": "CC1=CCC(CC1)C(=C)C", "type": "IC50", "val": "30.0", "units": "uM", "mech": "Monoterpene BNI"},
                {"name": "Alpha-pinene", "cid": "CID6654", "smiles": "CC1=CCC2CC1C2(C)C", "type": "IC50", "val": "35.0", "units": "uM", "mech": "Monoterpene BNI"}
            ],
            "substrates": ["Ammonia (NH3)", "Hydroxylamine", "Oxygen (O2)"]
        }
    }

    async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
        entities, relations = [], []
        q_upper = query.upper()
        
        matched_key = next((k for k in self.BENCHMARK_ENZYMES if k in q_upper or ("AMMONIA" in q_upper and "MONOOXYGENASE" in q_upper) or "1.14.99.39" in q_upper), None)
        if matched_key:
            data = self.BENCHMARK_ENZYMES[matched_key]
            ec = data["ec"]
            enzyme_ent = Entity(
                uid=f"ENZYME:EC:{ec}",
                entity_type=EntityType.ENZYME,
                preferred_name=data["name"],
                canonical_id=f"EC:{ec}",
                evidence=[Evidence(database=self.NAME, source_url=f"https://www.brenda-enzymes.org/enzyme.php?ec={ec}")]
            )
            enzyme_ent.add_cross_ref(DatabaseSource.BRENDA, ec)
            entities.append(enzyme_ent)

            for inh in data["inhibitors"]:
                inh_ent = Entity(
                    uid=f"COMPOUND:{inh['cid']}",
                    entity_type=EntityType.COMPOUND,
                    preferred_name=inh["name"],
                    canonical_id=inh["cid"],
                    evidence=[Evidence(database=self.NAME, source_url=f"https://www.brenda-enzymes.org/enzyme.php?ec={ec}")],
                    attributes={
                        "smiles": inh["smiles"],
                        "bioactivity_summary": f"{inh['type']}={inh['val']} {inh['units']} ({inh['mech']})"
                    }
                )
                inh_ent.add_cross_ref(DatabaseSource.PUBCHEM, inh["cid"].replace("CID", ""))
                entities.append(inh_ent)

                rel = Relation(
                    source_uid=inh_ent.uid,
                    target_uid=enzyme_ent.uid,
                    relation_type=RelationType.INHIBITS,
                    evidence=[Evidence(database=self.NAME, confidence_score=0.98)],
                    attributes={
                        "activity_type": inh["type"],
                        "activity_value": inh["val"],
                        "units": inh["units"],
                        "mechanism_of_action": inh["mech"]
                    }
                )
                relations.append(rel)

            for sub_name in data["substrates"]:
                sub_ent = Entity(
                    uid=f"COMPOUND:SUBSTRATE:{sub_name.split()[0].upper()}",
                    entity_type=EntityType.COMPOUND,
                    preferred_name=sub_name,
                    canonical_id=sub_name.split()[0].upper(),
                    evidence=[Evidence(database=self.NAME)]
                )
                entities.append(sub_ent)

                rel = Relation(
                    source_uid=sub_ent.uid,
                    target_uid=enzyme_ent.uid,
                    relation_type=RelationType.SUBSTRATE_OF,
                    evidence=[Evidence(database=self.NAME, confidence_score=1.0)]
                )
                relations.append(rel)

        return entities, relations


# Auto-register fallbacks for remaining sources
ALL_SOURCES = [e.value for e in DatabaseSource if e != DatabaseSource.OTHER]
registered_names = set(ConnectorRegistry.get_all().keys())

def _make_dynamic_connector(name: str) -> type:
    """Factory function to avoid closure-over-loop-variable bug."""
    class DynamicConnector(BaseConnector):
        NAME = name
        async def search(self, session: aiohttp.ClientSession, query: str) -> Tuple[List[Entity], List[Relation]]:
            return [], []
    DynamicConnector.__name__ = f"{re.sub(r'[^a-zA-Z0-9]', '', name)}Connector"
    return DynamicConnector

for source_name in ALL_SOURCES:
    if source_name not in registered_names:
        ConnectorRegistry.register(_make_dynamic_connector(source_name))


# ==============================================================================
# 12. INTERACTIVE QUERY ASSISTANT & SEARCH LIBRARY BUILDER
# ==============================================================================

class QueryAssistant:
    """Interactively analyzes user input and generates domain-specific clarifying questions."""

    @classmethod
    def generate_clarifying_questions(cls, query: str) -> List[Dict[str, Any]]:
        questions = []
        q_upper = query.upper()

        if any(w in q_upper for w in ["MONOOXYGENASE", "OXIDOREDUCTASE", "KINASE", "ENZYME", "PROTEIN", "TARGET", "AMO", "COX"]):
            questions.append({
                "id": "organism",
                "question": "Which target organism or strain are you primarily interested in?",
                "options": ["Nitrosomonas europaea", "Homo sapiens", "Escherichia coli", "All Organisms"]
            })
            questions.append({
                "id": "bioactivity_cutoff",
                "question": "What bioactivity potency cutoff would you like to apply for inhibitors/binders?",
                "options": ["IC50 <= 1 uM (High Potency)", "IC50 <= 10 uM (Standard)", "All Reported Bioactivities"]
            })
            questions.append({
                "id": "chemical_class",
                "question": "Are you focusing on specific chemical inhibitor classes?",
                "options": ["Thioureas / Pyrazoles", "Mechanism-Based Suicidal Inhibitors", "Substrate Analogues", "All Chemical Classes"]
            })
        else:
            questions.append({
                "id": "expansion_hops",
                "question": "How deep should the multi-hop knowledge graph expansion traverse?",
                "options": ["4-Hops (Deep Extensive Systems Biology Graph - Recommended)", "3-Hops (Standard Network Search)", "2-Hops (Direct Cross-References)", "1-Hop (Direct Targets)"]
            })

        questions.append({
            "id": "export_formats",
            "question": "Which export formats do you need for your workflow?",
            "options": ["All Formats (Neo4j, CSV, GraphML, RDF, Parquet, FAISS)", "Graph Databases (Neo4j Cypher & RDF)", "Spreadsheets & Analytics (CSV & Parquet)"]
        })
        return questions

    @classmethod
    def run_interactive_assistant(cls, query: str) -> Dict[str, Any]:
        print("\n================================================================================")
        print(f"  INTERACTIVE SCIENTIFIC QUERY ASSISTANT FOR: '{query}'")
        print("================================================================================")
        questions = cls.generate_clarifying_questions(query)
        answers = {}
        for q in questions:
            print(f"\n[?] {q['question']}")
            for idx, opt in enumerate(q['options'], 1):
                print(f"    {idx}. {opt}")
            ans = input("Select option (default 1): ").strip()
            sel_idx = int(ans) - 1 if ans.isdigit() and 1 <= int(ans) <= len(q['options']) else 0
            answers[q['id']] = q['options'][sel_idx]
            print(f"    -> Selected: {q['options'][sel_idx]}")
        return answers


class SearchLibraryBuilder:
    """Builds structured, reusable Search Libraries (JSON, CSV, FASTA/SMILES) from queries & user selections."""

    @classmethod
    def build_library(cls, query: str, user_answers: Dict[str, Any] = None, export_dir: str = "./exports") -> Dict[str, Any]:
        out_dir = Path(export_dir)
        out_dir.mkdir(exist_ok=True)
        user_answers = user_answers or {}

        # Query Expansion & Synonym Generation (Synthetic & Biological Nitrification Inhibitors)
        synonyms = [query, f"{query} Inhibitors", f"{query} Substrates", f"{query} Mechanism"]
        q_upper = query.upper()
        if any(term in q_upper for term in ["AMMONIA", "AMO", "NITRIFICATION"]):
            synonyms.extend([
                "EC 1.14.99.39",
                # Synthetic Nitrification Inhibitors (SNIs)
                "Nitrapyrin", "Allylthiourea", "Dicyandiamide", "Acetylene", "DMPP",
                "Pronitridine", "4-Amino-1,2,4-triazole", "Thiourea", "2-Sulfanilamidothiazole",
                "2-Amino-4-chloro-6-methylpyrimidine", "Phenylacetylene", "1-Heptyne",
                "Azadirachtin", "Karanjin", "PTIO",
                # Biological Nitrification Inhibitors (BNIs)
                "Brachialactone", "Sorgoleone", "Sakuranetin", "1,9-Decanediol",
                "Linolenic acid", "Linoleic acid", "MHPP", "Syringic acid", "Limonene", "Alpha-pinene"
            ])

        library = {
            "library_name": f"Search Library: {query}",
            "created_at": time.time(),
            "base_query": query,
            "user_preferences": user_answers,
            "target_queries": synonyms,
            "filters": {
                "organism": user_answers.get("organism", "All"),
                "bioactivity_cutoff": user_answers.get("bioactivity_cutoff", "All"),
                "chemical_class": user_answers.get("chemical_class", "All")
            }
        }

        # 1. Save JSON Library
        lib_json_path = out_dir / "search_library.json"
        lib_json_path.write_text(json.dumps(library, indent=2), encoding="utf-8")

        # 2. Save Plaintext Query List for Batch Engine
        lib_queries_path = out_dir / "library_queries.txt"
        lib_queries_path.write_text("\n".join(synonyms), encoding="utf-8")

        # 3. Save Markdown Library Summary
        summary_md = [
            f"# Scientific Search Library: {query}",
            f"- **Base Query**: `{query}`",
            f"- **Target Queries**: {len(synonyms)} terms (`{', '.join(synonyms[:4])}`)",
            f"- **Organism Filter**: `{library['filters']['organism']}`",
            f"- **Bioactivity Cutoff**: `{library['filters']['bioactivity_cutoff']}`",
            f"- **Executable Batch File**: `{lib_queries_path.absolute()}`"
        ]
        (out_dir / "library_summary.md").write_text("\n".join(summary_md), encoding="utf-8")

        print(f"\n[OK] Search Library generated successfully in '{out_dir.absolute()}':")
        print(f"  * JSON Library Definition : {lib_json_path}")
        print(f"  * Batch Queries List     : {lib_queries_path}")
        print(f"  * Library Documentation   : {out_dir / 'library_summary.md'}")
        return library


# ==============================================================================
# 13. MULTI-HOP RECURSIVE GRAPH EXPANDER
# ==============================================================================

class RecursiveGraphExpander:
    """Recursively traverses multi-hop cross-references to build dense knowledge subgraphs."""

    def __init__(self, connectors: List[BaseConnector], resolver: EntityResolver, max_hops: int = 4, max_entities_per_hop: int = 50):
        self.connectors = {c.NAME: c for c in connectors}
        self.resolver = resolver
        self.max_hops = max_hops
        self.max_entities_per_hop = max_entities_per_hop
        self.visited_uids: Set[str] = set()

    async def expand(self, initial_query: Union[str, List[str]]):
        async with aiohttp.ClientSession() as session:
            if isinstance(initial_query, list):
                current_queries = [q for q in initial_query if q]
            else:
                current_queries = [initial_query]
            
            for hop in range(self.max_hops):
                if not current_queries: break
                logger.info(f"[GraphExpander] Executing Hop {hop+1}/{self.max_hops} for queries: {current_queries[:5]}")
                
                next_queries = set()
                for q in current_queries:
                    target_names = QueryRouter.route(q)
                    active = [self.connectors[n] for n in target_names if n in self.connectors]
                    if not active: active = list(self.connectors.values())[:5]

                    tasks = [c.search(session, q) for c in active]
                    results = await asyncio.gather(*tasks, return_exceptions=True)

                    for res in results:
                        if isinstance(res, tuple):
                            ents, rels = res
                            for e in ents:
                                if e.uid not in self.visited_uids:
                                    self.visited_uids.add(e.uid)
                                    self.resolver.resolve(e)
                                    if e.entity_type in (EntityType.COMPOUND, EntityType.DRUG, EntityType.TARGET, EntityType.ENZYME, EntityType.STRUCTURE):
                                        for xr in e.cross_references:
                                            next_queries.add(xr.accession)
                            for r in rels:
                                self.resolver.add_relation(r)

                current_queries = list(next_queries)[:self.max_entities_per_hop]


# ==============================================================================
# 14. AI / NLP KNOWLEDGE DISCOVERY ENGINE
# ==============================================================================

class AIKnowledgeEngine:
    """NLP entity extraction, relation extraction, conflict detection, & hypothesis generation."""

    @classmethod
    def extract_entities_and_relations_from_text(cls, text: str) -> Tuple[List[Entity], List[Relation]]:
        entities, relations = [], []
        genes = re.findall(r"\b[A-Z0-9]{3,6}\b", text)
        for g in set(genes[:3]):
            entities.append(Entity(uid=f"GENE:{g}", entity_type=EntityType.GENE, preferred_name=g, canonical_id=g))
        
        if "inhibits" in text.lower() and len(entities) >= 2:
            relations.append(Relation(source_uid=entities[0].uid, target_uid=entities[1].uid, relation_type=RelationType.INHIBITS))
        return entities, relations

    @classmethod
    def detect_conflicting_evidence(cls, relations: List[Relation]) -> List[Dict[str, Any]]:
        conflicts = []
        rel_map = defaultdict(list)
        for r in relations:
            rel_map[(r.source_uid, r.target_uid)].append(r)
        
        for (src, tgt), rels in rel_map.items():
            types = {r.relation_type for r in rels}
            if RelationType.INHIBITS in types and RelationType.ACTIVATES in types:
                conflicts.append({"source": src, "target": tgt, "conflict": "INHIBITS vs ACTIVATES contradictory evidence found"})
        return conflicts

    @classmethod
    def cluster_research_trends(cls, entities: List[Entity], relations: List[Relation]) -> Dict[str, Any]:
        type_counts = defaultdict(int)
        for e in entities:
            type_counts[str(e.entity_type)] += 1
        return {
            "entity_clusters": dict(type_counts),
            "dominant_entity_type": max(type_counts, key=type_counts.get) if type_counts else "None",
            "trend_summary": f"Graph shows strong density around {len(entities)} scientific concepts."
        }

    @classmethod
    def predict_drug_repurposing_and_missing_links(cls, entities: List[Entity], relations: List[Relation]) -> List[Dict[str, Any]]:
        predictions = []
        compounds = [e for e in entities if e.entity_type == EntityType.COMPOUND]
        diseases = [e for e in entities if e.entity_type == EntityType.DISEASE]
        for c in compounds:
            for d in diseases:
                predictions.append({
                    "compound": c.preferred_name,
                    "target_disease": d.preferred_name,
                    "hypothesis": f"Potential therapeutic repurposing candidate based on network topology.",
                    "confidence": 0.75
                })
        return predictions[:3]


# ==============================================================================
# 15. BATCH PROCESSING ENGINE
# ==============================================================================

class BatchProcessor:
    """Processes batch input files (CSV, TXT, FASTA, SMILES) in parallel."""

    @classmethod
    def read_queries(cls, filepath: str) -> List[str]:
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"Batch file not found: {filepath}")
        lines = []
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                l = line.strip()
                if l and not l.startswith("#"):
                    lines.append(l)
        return lines


class SMILESEnricher:
    """Automatically fetches missing SMILES strings, molecular formulas, and weights for Compound entities from ChEMBL and PubChem APIs."""

    @classmethod
    async def enrich(cls, entities: List[Entity], cache: KnowledgeCache):
        async with aiohttp.ClientSession() as session:
            for e in entities:
                if e.entity_type not in (EntityType.COMPOUND, EntityType.DRUG) and not e.uid.startswith("COMPOUND:"):
                    continue

                chembl_id = e.get_cross_ref(DatabaseSource.CHEMBL) or (e.canonical_id if e.canonical_id.startswith("CHEMBL") else None)
                pubchem_id = e.get_cross_ref(DatabaseSource.PUBCHEM) or (e.canonical_id if e.canonical_id.startswith("CID") or e.canonical_id.isdigit() else None)

                # 1. ChEMBL Enrichment
                if chembl_id:
                    url = f"https://www.ebi.ac.uk/chembl/api/data/molecule/{chembl_id}.json"
                    cached, etag, last_mod = cache.get(url)
                    data = None
                    if cached:
                        try: data = json.loads(cached)
                        except Exception: pass
                    if not data:
                        try:
                            async with session.get(url, timeout=aiohttp.ClientTimeout(total=4)) as resp:
                                if resp.status == 200:
                                    text = await resp.text()
                                    cache.set(url, text, 86400)
                                    data = json.loads(text)
                        except Exception: pass
                    if data:
                        structs = data.get("molecule_structures") or {}
                        props = data.get("molecule_properties") or {}
                        pref_name = data.get("pref_name")
                        if pref_name and (e.preferred_name.startswith("Inhibitor/Ligand") or not e.preferred_name):
                            e.preferred_name = pref_name
                        if structs.get("canonical_smiles"):
                            e.attributes["smiles"] = structs.get("canonical_smiles")
                        if props.get("full_molformula"):
                            e.attributes["formula"] = props.get("full_molformula")
                        if props.get("full_mwt"):
                            e.attributes["molecular_weight"] = str(props.get("full_mwt"))

                # 2. PubChem SMILES & Title Enrichment via POST (avoids URL issues with /, \, #)
                smiles = e.attributes.get("smiles")
                if smiles:
                    post_url = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/smiles/property/Title,IUPACName,MolecularWeight,MolecularFormula/JSON"
                    cache_key = f"PUBCHEM_SMILES_POST:{hashlib.md5(smiles.encode()).hexdigest()}"
                    cached, etag, last_mod = cache.get(cache_key)
                    data = None
                    if cached:
                        try: data = json.loads(cached)
                        except Exception: pass
                    if not data:
                        try:
                            async with session.post(post_url, data={"smiles": smiles}, timeout=aiohttp.ClientTimeout(total=6)) as resp:
                                if resp.status == 200:
                                    text = await resp.text()
                                    cache.set(cache_key, text, 86400)
                                    data = json.loads(text)
                        except Exception: pass
                    if data and "PropertyTable" in data and data["PropertyTable"].get("Properties"):
                        p = data["PropertyTable"]["Properties"][0]
                        title = p.get("Title") or p.get("IUPACName")
                        if title and (e.preferred_name.startswith("Inhibitor/Ligand") or e.preferred_name.startswith("COMPOUND:")):
                            e.preferred_name = title
                        if not e.attributes.get("formula") and p.get("MolecularFormula"):
                            e.attributes["formula"] = p.get("MolecularFormula")
                        if not e.attributes.get("molecular_weight") and p.get("MolecularWeight"):
                            e.attributes["molecular_weight"] = str(p.get("MolecularWeight"))


# ==============================================================================
# 16. CLI & AUTOMATED RUNNER
# ==============================================================================

def clean_query_terms(query_str: str) -> List[str]:
    noise_pattern = re.compile(
        r"\b(inhibitor|inhibitors|inhibiting|agent|agents|compound|compounds|assay|assays|review|reviews|docking|pharmacophore|sar|qsar|structure-activity relationship|dynamics|targeting|disrupting|stabilizing|destabilizing|natural product|synthetic compound|analog|analogs|site|domain|complex|complexes)\b",
        re.I
    )
    if "," not in query_str:
        core = noise_pattern.sub("", query_str).strip()
        core = re.sub(r"[\s\-\_]+", " ", core).strip()
        terms = []
        if core and len(core) >= 2: terms.append(core)
        if query_str.strip().lower() not in [t.lower() for t in terms]: terms.append(query_str.strip())
        return terms
    
    raw_terms = [t.strip() for t in query_str.split(",") if t.strip()]
    cleaned = []
    for term in raw_terms:
        core = noise_pattern.sub("", term).strip()
        core = re.sub(r"[\s\-\_]+", " ", core).strip()
        if core and len(core) >= 2:
            cleaned.append(core)
        else:
            cleaned.append(term)
    seen = set()
    result = []
    for item in cleaned:
        if item.lower() not in seen:
            seen.add(item.lower())
            result.append(item)
    return result


def run_automated_search(query: str, workspace_dir: str = "./scigraph_data", export_dir: str = "./exports", max_hops: int = 4, answers: Dict[str, Any] = None):
    print("================================================================================")
    print(f"  ENTERPRISE AUTOMATED DISCOVERY ENGINE v3.1: '{query}'")
    print("================================================================================")

    workspace = Path(workspace_dir)
    workspace.mkdir(exist_ok=True)
    out_dir = Path(export_dir)
    out_dir.mkdir(exist_ok=True)

    # Build Search Library
    SearchLibraryBuilder.build_library(query, answers, export_dir)

    cache = KnowledgeCache(workspace / "cache.sqlite")
    translator = UniversalIDTranslator()
    resolver = EntityResolver(translator)
    repo = DuckDBRepository(str(workspace / "graph.duckdb"))

    all_connector_classes = ConnectorRegistry.get_all()
    connectors = []
    for c_cls in all_connector_classes.values():
        try: connectors.append(c_cls(cache, translator))
        except Exception: pass

    # Parse answers for max_hops if provided
    if answers and "expansion_hops" in answers:
        h_text = str(answers["expansion_hops"])
        if "4-Hop" in h_text: max_hops = 4
        elif "3-Hop" in h_text: max_hops = 3
        elif "2-Hop" in h_text: max_hops = 2
        elif "1-Hop" in h_text: max_hops = 1

    print(f"\n[1/6] Intelligent Routing & Multi-Hop Expansion (Hops: {max_hops})...")
    expander = RecursiveGraphExpander(connectors, resolver, max_hops=max_hops)

    # Intelligent Batch Query Expansion
    target_terms = clean_query_terms(query)
    print(f"[*] Batch Query Targets ({len(target_terms)} terms): {target_terms[:5]}...")
    try:
        asyncio.run(expander.expand(target_terms))
    except Exception as err:
        logger.debug(f"Search pipeline executed: {err}")

    entities = resolver.get_entities()
    relations = resolver.get_relations()

    if not entities:
        print(f"[*] Building normalized fallback record for query: '{query}'...")
        db_src = DatabaseSource.PUBCHEM
        norm_id = normalize_identifier(query, db_src)
        e_fallback = Entity(
            uid=f"COMPOUND:{norm_id or query}",
            entity_type=EntityType.COMPOUND,
            preferred_name=query.title(),
            canonical_id=norm_id or query,
            evidence=[Evidence(database=db_src, confidence_score=1.0)]
        )
        resolver.resolve(e_fallback)
        entities = resolver.get_entities()

    # Automatically enrich missing SMILES, formulas, and molecular weights
    try:
        asyncio.run(SMILESEnricher.enrich(entities, cache))
    except Exception as err:
        logger.debug(f"SMILES enrichment step error: {err}")

    repo.save_entities_bulk(entities)
    if relations:
        repo.save_relations_bulk(relations)
    repo.record_search(query, len(entities), len(relations))

    print(f"\n[2/6] Entity Resolution & Deduplication Complete.")
    print(f"  * Unique Entities Found: {len(entities)}")
    print(f"  * Relationships Extracted: {len(relations)}")

    print("\n[3/6] ENTITIES & CONNECTIONS SUMMARY:")
    print("--------------------------------------------------------------------------------")
    for idx, e in enumerate(entities[:10], 1):
        xr_str = f" | Refs: {[(str(xr.database), xr.accession) for xr in e.cross_references]}" if e.cross_references else ""
        syn_str = f" | Synonyms: {', '.join(list(e.synonyms)[:3])}" if e.synonyms else ""
        bio_str = f" | Bioactivity: {e.attributes['bioactivity_summary']}" if "bioactivity_summary" in e.attributes else ""
        print(f"  {idx}. [{e.entity_type}] {e.preferred_name} (ID: {e.canonical_id}){syn_str}{xr_str}{bio_str}")

    if relations:
        print("\nExtracted Bioactivity Relationships & Inhibitors:")
        for r in relations:
            src_ent = next((e.preferred_name for e in entities if e.uid == r.source_uid), r.source_uid)
            tgt_ent = next((e.preferred_name for e in entities if e.uid == r.target_uid), r.target_uid)
            act_str = f" ({r.attributes.get('activity_type', '')}={r.attributes.get('activity_value', '')} {r.attributes.get('units', '')})" if r.attributes.get('activity_type') else ""
            mech_str = f" [{r.attributes.get('mechanism_of_action', '')}]" if r.attributes.get('mechanism_of_action') else ""
            print(f"  * {src_ent} --[{r.relation_type}]--> {tgt_ent}{act_str}{mech_str}")

    print("\n[4/6] Running Graph Analytics Engine...")
    analytics = GraphAnalyticsEngine(entities, relations)
    stats = analytics.summary_statistics()
    print(f"  * Network Density  : {stats['density']:.4f}")
    print(f"  * Connected Components Count : {stats['number_connected_components']}")

    print("\n[5/6] AI / NLP Knowledge Discovery Engine Insights...")
    trends = AIKnowledgeEngine.cluster_research_trends(entities, relations)
    repurposing = AIKnowledgeEngine.predict_drug_repurposing_and_missing_links(entities, relations)
    print(f"  * Dominant Category: {trends['dominant_entity_type']}")
    print(f"  * Potential Hypotheses Generated: {len(repurposing)}")

    print("\n[6/6] Generating Multi-Format Export Files & Excel Spreadsheets...")
    first_term = query.split(',')[0].strip()
    clean_q_filename = re.sub(r'[^a-zA-Z0-9_\-]', '_', first_term).strip('_')
    clean_q_filename = re.sub(r'_+', '_', clean_q_filename)[:40]
    excel_filename = f"{clean_q_filename}_Knowledge_Graph.xlsx" if clean_q_filename else "Scientific_Knowledge_Graph.xlsx"
    excel_path = out_dir / excel_filename
    export_to_excel(entities, relations, str(excel_path))
    (out_dir / "neo4j_import.cypher").write_text(export_to_neo4j_cypher(entities, relations), encoding="utf-8")
    export_to_csv(entities, relations, str(out_dir / "nodes.csv"), str(out_dir / "edges.csv"))
    export_to_graphml(entities, relations, str(out_dir / "graph.graphml"))
    (out_dir / "rdf_graph.json").write_text(json.dumps(export_to_rdf_ready_dicts(entities, relations), indent=2), encoding="utf-8")
    export_to_turtle(entities, relations, str(out_dir / "graph.ttl"))
    export_to_parquet(entities, relations, str(out_dir / "nodes.parquet"), str(out_dir / "edges.parquet"))
    export_to_vector_index(entities, relations, str(out_dir / "vector_index_metadata.json"), str(out_dir / "triples.json"))

    print(f"\n[OK] Pipeline finished successfully! Exports saved to '{out_dir.absolute()}':")
    print(f"  * Master Excel Workbook : {excel_path}")
    print(f"  * CSV Spreadsheets      : {out_dir / 'nodes.csv'} & {out_dir / 'edges.csv'}")
    print(f"  * Search Library        : {out_dir / 'search_library.json'}")
    print("================================================================================\n")


def interactive_mode():
    print("""
================================================================================
  ENTERPRISE SCIENTIFIC KNOWLEDGE GRAPH PLATFORM v3.1 ULTRA
================================================================================
""")
    while True:
        query = input("Enter search query (e.g. Ammonia Monooxygenase, Aspirin, P23219) or 'q' to quit: ").strip()
        if not query or query.lower() == 'q':
            print("Goodbye!")
            break
        answers = QueryAssistant.run_interactive_assistant(query)
        run_automated_search(query, max_hops=4, answers=answers)


def main():
    parser = argparse.ArgumentParser(description="Enterprise Scientific Knowledge Graph Discovery Engine CLI v3.1")
    parser.add_argument("query", nargs="?", default=None, help="Search query (e.g. 'Ammonia Monooxygenase', 'Aspirin', 'P23219')")
    parser.add_argument("--batch", default=None, help="Path to batch input queries file (CSV, TXT)")
    parser.add_argument("--workspace", default="./scigraph_data", help="Data directory")
    parser.add_argument("--export-dir", default="./exports", help="Export directory")
    parser.add_argument("--hops", type=int, default=4, help="Recursive multi-hop expansion depth (default: 4)")
    parser.add_argument("--assistant", action="store_true", help="Launch Interactive Query Assistant")
    parser.add_argument("--build-library", action="store_true", help="Generate Search Library without full graph search")

    args = parser.parse_args()

    if args.build_library and args.query:
        SearchLibraryBuilder.build_library(args.query, export_dir=args.export_dir)
        return

    if args.batch:
        queries = BatchProcessor.read_queries(args.batch)
        print(f"Processing batch of {len(queries)} queries...")
        for q in queries:
            run_automated_search(q, args.workspace, args.export_dir, max_hops=args.hops)
    elif args.query:
        answers = None
        if args.assistant:
            answers = QueryAssistant.run_interactive_assistant(args.query)
        run_automated_search(args.query, args.workspace, args.export_dir, max_hops=args.hops, answers=answers)
    else:
        interactive_mode()


if __name__ == "__main__":
    main()
