#!/usr/bin/env python3
"""
Enrichment Pipeline: API-powered data enrichment for chemical-data-extractor exports.

Fetches SMILES, 2D structure images, molecular properties, and publication details,
then writes everything to a beautifully formatted multi-sheet Excel workbook.

Usage:
    python3 enrich_exports.py                           # Uses ./exports/
    python3 enrich_exports.py --export-dir ./my_output   # Custom directory
"""
import argparse
import csv
import io
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from collections import defaultdict
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration (overridable via --export-dir) ───────────────────────────
EXPORT_DIR = "exports"
NODES_FILE: str = ""
EDGES_FILE: str = ""
OUTPUT_XLSX: str = ""
IMG_DIR: str = ""


def set_export_dir(path: str):
    """Configure all paths to point at the given export directory."""
    global EXPORT_DIR, NODES_FILE, EDGES_FILE, OUTPUT_XLSX, IMG_DIR
    EXPORT_DIR = path
    NODES_FILE = os.path.join(path, "nodes.csv")
    EDGES_FILE = os.path.join(path, "edges.csv")
    OUTPUT_XLSX = os.path.join(path, "enriched_data.xlsx")
    IMG_DIR = os.path.join(path, "structure_images")
    os.makedirs(IMG_DIR, exist_ok=True)

# PubChem API
PUBCHEM_BASE = "https://pubchem.ncbi.nlm.nih.gov/rest/pug"
PUBCHEM_IMG = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound"

# CrossRef API
CROSSREF_BASE = "https://api.crossref.org/works"

# Cache
PUBCHEM_CACHE = {}
PUBCHEM_IMG_CACHE = set()

# ── Styling helpers ─────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT = Font(name="Calibri", size=10)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
INHIBIT_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
BIND_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data(ws, nrows, ncols, alt_rows=True):
    for row in range(2, nrows + 2):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
            if alt_rows and row % 2 == 0:
                cell.fill = ALT_FILL


def auto_width(ws, ncols, max_width=60):
    for col in range(1, ncols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 3


# ── API Helpers ─────────────────────────────────────────────────────────────

def _pubchem_cid_properties(cid: int) -> Optional[dict]:
    """Get compound properties from PubChem by CID.
    Uses /property/ endpoint first, falls back to full /JSON record if SMILES missing."""
    result = None

    # Try /property/ endpoint first (fast)
    url = f"{PUBCHEM_BASE}/compound/cid/{cid}/property/MolecularFormula,MolecularWeight,CanonicalSMILES,IsomericSMILES,InChI,InChIKey,IUPACName,Title/JSON"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "ChemicalDataExtractor/1.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
            if "PropertyTable" in data and "Properties" in data["PropertyTable"]:
                props = data["PropertyTable"]["Properties"][0]
                smiles = props.get("CanonicalSMILES") or props.get("IsomericSMILES") or ""
                result = {
                    "smiles": smiles,
                    "formula": props.get("MolecularFormula", ""),
                    "mw": str(props.get("MolecularWeight", "")),
                    "inchi": props.get("InChI", ""),
                    "inchikey": props.get("InChIKey", ""),
                    "iupac": props.get("IUPACName", ""),
                    "cid": str(props.get("CID", cid)),
                    "title": props.get("Title", ""),
                }
                if smiles:  # Got SMILES, return immediately
                    return result
    except Exception:
        pass

    # Fallback: use full JSON record (has SMILES under props, not property endpoint)
    if not result or not result.get("smiles"):
        url2 = f"{PUBCHEM_BASE}/compound/cid/{cid}/JSON"
        try:
            req2 = urllib.request.Request(url2, headers={"User-Agent": "ChemicalDataExtractor/1.0"})
            with urllib.request.urlopen(req2, timeout=15) as resp2:
                data2 = json.loads(resp2.read())
                pc = data2.get("PC_Compounds", [{}])[0]
                props2 = pc.get("props", [])

                smiles = result["smiles"] if result else ""
                formula = result["formula"] if result else ""
                mw = result["mw"] if result else ""
                inchi = result["inchi"] if result else ""
                inchikey = result["inchikey"] if result else ""
                iupac = result["iupac"] if result else ""

                for p in props2:
                    urn = p.get("urn", {})
                    label = urn.get("label", "")
                    name = urn.get("name", "") or ""
                    sval = p.get("value", {}).get("sval", "")
                    if sval:
                        if "SMILES" in label.upper() and not smiles:
                            smiles = sval
                        elif label == "Molecular Formula" and not formula:
                            formula = sval
                        elif label == "Molecular Weight" and not mw:
                            mw = sval
                        elif label == "InChI" and not inchi:
                            inchi = sval
                        elif label == "InChIKey" and not inchikey:
                            inchikey = sval
                        elif label == "IUPAC Name" and not iupac:
                            iupac = sval

                # Prefer connectivity SMILES over absolute
                for p in props2:
                    urn = p.get("urn", {})
                    label = urn.get("label", "")
                    name = urn.get("name", "") or ""
                    sval = p.get("value", {}).get("sval", "")
                    if sval and label == "SMILES" and "Connectivity" in name:
                        smiles = sval
                        break

                result = {
                    "smiles": smiles or "",
                    "formula": formula or "",
                    "mw": mw or "",
                    "inchi": inchi or "",
                    "inchikey": inchikey or "",
                    "iupac": iupac or "",
                    "cid": str(cid),
                    "title": "",
                }
                return result
        except Exception:
            pass

    return result


def _pubchem_name_to_cid(name: str) -> Optional[int]:
    """Resolve a compound name to a PubChem CID. Uses /cids/JSON endpoint which is more reliable."""
    encoded = urllib.parse.quote(name)
    url = f"{PUBCHEM_BASE}/compound/name/{encoded}/cids/JSON"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "ChemicalDataExtractor/1.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
            if "IdentifierList" in data and "CID" in data["IdentifierList"]:
                return data["IdentifierList"]["CID"][0]
    except Exception:
        pass
    return None


def pubchem_smiles(name: str) -> Optional[dict]:
    """Look up compound properties from PubChem by name.
    Uses two-step pipeline: name → CID → properties (more reliable than name→properties directly).
    Falls back to cleaning the name and retrying if initial lookup fails."""
    if name in PUBCHEM_CACHE:
        return PUBCHEM_CACHE[name]

    # Collect all name variations to try
    names_to_try = [name]

    # Also try cleaning various prefixes
    clean = name.strip()
    for prefix in ["BRENDA_INHIBITOR:", "COMPOUND:", "CHEMBL"]:
        if clean.upper().startswith(prefix.upper()):
            clean = clean[len(prefix):].strip()
            names_to_try.append(clean)
            break

    # For names with abbreviations like "Dicyandiamide (DCD)", also try the bare name
    if "(" in clean and ")" in clean:
        bare = clean.split("(")[0].strip()
        if bare and bare != clean:
            names_to_try.append(bare)

    # Try each name variation
    for n in names_to_try:
        cid = _pubchem_name_to_cid(n)
        if cid:
            result = _pubchem_cid_properties(cid)
            if result:
                PUBCHEM_CACHE[name] = result
                time.sleep(0.3)
                return result
        time.sleep(0.2)

    # Final fallback: try a broad synonym search
    try:
        encoded = urllib.parse.quote(names_to_try[0])
        url = f"{PUBCHEM_BASE}/compound/name/{encoded}/synonyms/JSON"
        req = urllib.request.Request(url, headers={"User-Agent": "ChemicalDataExtractor/1.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
            if "InformationList" in data and "Information" in data["InformationList"]:
                for info in data["InformationList"]["Information"]:
                    cid = info.get("CID")
                    if cid:
                        result = _pubchem_cid_properties(cid)
                        if result:
                            PUBCHEM_CACHE[name] = result
                            time.sleep(0.3)
                            return result
    except Exception:
        pass

    PUBCHEM_CACHE[name] = None
    return None


def download_pubchem_image(smiles: str, cid: str = "", filename: str = "") -> Optional[str]:
    """Download 2D structure image from PubChem using CID (preferred) or SMILES."""
    if not smiles and not cid:
        return None
    if not filename:
        return None

    filepath = os.path.join(IMG_DIR, filename)
    if os.path.exists(filepath) and os.path.getsize(filepath) > 100:
        return filepath

    # Try CID-based URL first (most reliable)
    if cid:
        url = f"{PUBCHEM_IMG}/cid/{cid}/PNG?record_type=2d&image_size=200x200"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "ChemicalDataExtractor/1.0"})
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = resp.read()
                if len(data) > 100:
                    with open(filepath, "wb") as f:
                        f.write(data)
                    return filepath
        except Exception:
            pass

    # Fallback to SMILES-based URL
    if smiles:
        url = f"{PUBCHEM_IMG}/smiles/{urllib.parse.quote(smiles)}/PNG?record_type=2d&image_size=200x200"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "ChemicalDataExtractor/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = resp.read()
                if len(data) > 100:
                    with open(filepath, "wb") as f:
                        f.write(data)
                    return filepath
        except Exception:
            pass

    return None


def crossref_lookup(doi: str) -> Optional[dict]:
    """Look up publication metadata from CrossRef API by DOI."""
    doi_clean = doi.replace("https://doi.org/", "").replace("http://dx.doi.org/", "").strip()
    if not doi_clean:
        return None

    url = f"{CROSSREF_BASE}/{doi_clean}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "ChemicalDataExtractor/1.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
            msg = data.get("message", {})

            authors = []
            for a in msg.get("author", []):
                given = a.get("given", "")
                family = a.get("family", "")
                if given and family:
                    authors.append(f"{given} {family}")
                elif family:
                    authors.append(family)

            abstract = msg.get("abstract", "")
            # Strip HTML from abstract
            if abstract:
                abstract = re.sub(r"<[^>]+>", "", abstract)
                abstract = abstract.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")

            result = {
                "title": msg.get("title", [""])[0] if msg.get("title") else "",
                "authors": "; ".join(authors[:20]),
                "journal": msg.get("container-title", [""])[0] if msg.get("container-title") else "",
                "year": str(msg.get("published-print", {}).get("date-parts", [[None]])[0][0] or
                            msg.get("published-online", {}).get("date-parts", [[None]])[0][0] or ""),
                "volume": msg.get("volume", ""),
                "issue": msg.get("issue", ""),
                "pages": msg.get("page", ""),
                "abstract": abstract[:2000] if abstract else "",
                "url": msg.get("URL", f"https://doi.org/{doi_clean}"),
                "doi": doi_clean,
                "publisher": msg.get("publisher", ""),
                "type": msg.get("type", ""),
                "is_referenced_by_count": str(msg.get("is-referenced-by-count", 0)),
            }
            time.sleep(0.2)
            return result
    except Exception:
        return None


# ── Data Loading ────────────────────────────────────────────────────────────

def load_nodes():
    """Load all nodes from CSV."""
    nodes = {}
    with open(NODES_FILE, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            uid = row.get("uid:ID", "")
            nodes[uid] = row
    return nodes


def load_edges():
    """Load all edges from CSV."""
    edges = []
    with open(EDGES_FILE, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            edges.append(row)
    return edges


# ── Enrichment Functions ────────────────────────────────────────────────────

def enrich_compounds(nodes, edges):
    """Enrich compound nodes with SMILES, properties from PubChem."""
    enriched = 0
    compound_nodes = {}
    compound_uids = set()

    # Find compound nodes
    for uid, node in nodes.items():
        label = node.get(":LABEL", "")
        if "COMPOUND" in label.upper():
            compound_nodes[uid] = node
            compound_uids.add(uid)

    print(f"  Found {len(compound_nodes)} compound nodes")

    for uid, node in compound_nodes.items():
        name = node.get("name", "").strip()
        existing_smiles = node.get("smiles", "").strip()

        if existing_smiles:
            node["_enriched_smiles"] = existing_smiles
            node["_pubchem_cid"] = node.get("canonical_id", "").replace("CID:", "")
            continue

        if not name:
            continue

        result = pubchem_smiles(name)
        if result:
            node["_enriched_smiles"] = result["smiles"]
            node["_enriched_formula"] = result["formula"]
            node["_enriched_mw"] = result["mw"]
            node["_enriched_inchi"] = result["inchi"]
            node["_enriched_inchikey"] = result["inchikey"]
            node["_enriched_iupac"] = result["iupac"]
            node["_pubchem_cid"] = result["cid"]
            enriched += 1

    print(f"  Enriched {enriched} compounds with PubChem data")
    return compound_nodes


def enrich_publications(nodes):
    """Enrich publication nodes with metadata from CrossRef."""
    enriched = 0
    pub_nodes = {}

    for uid, node in nodes.items():
        label = node.get(":LABEL", "")
        if "PUBLICATION" in label.upper() or "PUBLICATION" in label.upper():
            # Extract DOI
            doi = ""
            for field in ["canonical_id", "cross_references", "name", "uid:ID"]:
                val = node.get(field, "")
                if "doi.org" in val.lower() or val.startswith("10."):
                    m = re.search(r"10\.\d{4,}/[^\s,;]+", val)
                    if m:
                        doi = m.group(0)
                        break

            if doi:
                meta = crossref_lookup(doi)
                if meta:
                    node["_crossref_title"] = meta["title"]
                    node["_crossref_authors"] = meta["authors"]
                    node["_crossref_journal"] = meta["journal"]
                    node["_crossref_year"] = meta["year"]
                    node["_crossref_abstract"] = meta["abstract"]
                    node["_crossref_url"] = meta["url"]
                    node["_crossref_doi"] = meta["doi"]
                    node["_crossref_publisher"] = meta["publisher"]
                    node["_crossref_citations"] = meta["is_referenced_by_count"]
                    enriched += 1
                else:
                    # Store DOI even without crossref data
                    node["_crossref_doi"] = doi

            pub_nodes[uid] = node

    print(f"  Found {len(pub_nodes)} publications, enriched {enriched}")
    return pub_nodes


def download_structure_images(compound_nodes):
    """Download 2D structure images for compounds using CID (preferred) or SMILES."""
    downloaded = 0
    for uid, node in compound_nodes.items():
        smiles = node.get("_enriched_smiles", node.get("smiles", "")).strip()
        cid = node.get("_pubchem_cid", "").strip()

        # Must have a CID or SMILES
        if not smiles and not cid:
            continue

        name = node.get("name", uid)[:40]
        safe_name = re.sub(r"[^\w\-_]", "_", name)
        filename = f"{safe_name}_{uid[:12]}.png"

        if filename in PUBCHEM_IMG_CACHE:
            continue

        img_path = download_pubchem_image(smiles, cid, filename)
        if img_path:
            node["_structure_img"] = img_path
            PUBCHEM_IMG_CACHE.add(filename)
            downloaded += 1
            if downloaded % 10 == 0:
                print(f"    Downloaded {downloaded} images...")

    print(f"  Downloaded {downloaded} 2D structure images")
    return downloaded


# ── Excel Writing ───────────────────────────────────────────────────────────

def img_to_excel(ws, img_path, row, col):
    """Embed an image into an Excel cell."""
    if not img_path or not os.path.exists(img_path):
        return
    try:
        img = openpyxl.drawing.image.Image(img_path)
        img.width = 80
        img.height = 80
        img.anchor = f"{get_column_letter(col)}{row}"
        ws.add_image(img)
    except Exception:
        pass


def write_compounds_sheet(wb, compound_nodes, all_nodes):
    """Sheet 2: All compounds with SMILES, properties, and 2D structure images."""
    ws = wb.create_sheet("Compounds", 1)

    headers = [
        "Name", "SMILES", "Molecular Formula", "Molecular Weight",
        "InChI Key", "IUPAC Name", "PubChem CID", "Source DB",
        "Bioactivity Summary", "2D Structure"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    row = 2
    for uid, node in sorted(compound_nodes.items()):
        name = node.get("name", "")
        smiles = node.get("_enriched_smiles", node.get("smiles", ""))
        formula = node.get("_enriched_formula", node.get("formula", ""))
        mw = node.get("_enriched_mw", node.get("molecular_weight", ""))
        inchikey = node.get("_enriched_inchikey", "")
        iupac = node.get("_enriched_iupac", "")
        cid = node.get("_pubchem_cid", "")
        source = node.get(":LABEL", "").replace("COMPOUND:", "").replace("COMPOUND", "").strip()
        bio = node.get("bioactivity_summary", "")
        img_path = node.get("_structure_img", "")

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=smiles)
        ws.cell(row=row, column=3, value=formula)
        ws.cell(row=row, column=4, value=mw)
        ws.cell(row=row, column=5, value=inchikey)
        ws.cell(row=row, column=6, value=iupac)
        ws.cell(row=row, column=7, value=cid)
        ws.cell(row=row, column=8, value=source)
        ws.cell(row=row, column=9, value=bio)

        if img_path:
            img_to_excel(ws, img_path, row, 10)

        row += 1

    style_data(ws, row - 2, len(headers))
    ws.column_dimensions["J"].width = 22  # Image column
    auto_width(ws, len(headers) - 1)

    print(f"  Sheet 'Compounds': {row - 2} rows")
    return ws


def write_inhibitors_sheet(wb, nodes, edges):
    """Sheet 2: Inhibitor activities with compound → target mappings."""
    ws = wb.create_sheet("Inhibitor Activities")

    headers = [
        "Compound Name", "Compound SMILES", "Target Name", "Target ID",
        "Activity Type", "Activity Value", "Units", "pChEMBL Value",
        "Mechanism of Action", "Assay ID", "Relation Type", "Confidence"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    row = 2
    for e in edges:
        rel_type = e.get(":TYPE", "")
        if rel_type not in ("INHIBITS", "BINDS", "ACTIVE_AGAINST"):
            continue

        src_id = e.get(":START_ID", "")
        tgt_id = e.get(":END_ID", "")
        src_node = nodes.get(src_id, {})
        tgt_node = nodes.get(tgt_id, {})

        src_name = src_node.get("name", src_id)
        tgt_name = tgt_node.get("name", tgt_id)
        src_smiles = src_node.get("_enriched_smiles", src_node.get("smiles", ""))

        ws.cell(row=row, column=1, value=src_name)
        ws.cell(row=row, column=2, value=src_smiles)
        ws.cell(row=row, column=3, value=tgt_name)
        ws.cell(row=row, column=4, value=tgt_id)
        ws.cell(row=row, column=5, value=e.get("activity_type", ""))
        ws.cell(row=row, column=6, value=e.get("activity_value", ""))
        ws.cell(row=row, column=7, value=e.get("units", ""))
        ws.cell(row=row, column=8, value=e.get("pchembl_value", ""))
        ws.cell(row=row, column=9, value=e.get("mechanism_of_action", ""))
        ws.cell(row=row, column=10, value=e.get("assay_id", ""))
        ws.cell(row=row, column=11, value=rel_type)
        ws.cell(row=row, column=12, value=e.get("confidence", ""))

        # Color-code by relation type
        fill = INHIBIT_FILL if rel_type == "INHIBITS" else BIND_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = fill

        row += 1

    style_data(ws, row - 2, len(headers), alt_rows=False)
    auto_width(ws, len(headers))

    print(f"  Sheet 'Inhibitor Activities': {row - 2} rows")
    return ws


def write_publications_sheet(wb, pub_nodes):
    """Sheet 3: Publications with metadata from CrossRef."""
    ws = wb.create_sheet("Publications")

    headers = [
        "DOI", "Title", "Authors", "Journal", "Year", "Volume",
        "Pages", "Publisher", "Type", "Cited By Count", "Abstract",
        "URL"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    row = 2
    for uid, node in sorted(pub_nodes.items()):
        ws.cell(row=row, column=1, value=node.get("_crossref_doi", ""))
        ws.cell(row=row, column=2, value=node.get("_crossref_title", node.get("name", "")))
        ws.cell(row=row, column=3, value=node.get("_crossref_authors", ""))
        ws.cell(row=row, column=4, value=node.get("_crossref_journal", ""))
        ws.cell(row=row, column=5, value=node.get("_crossref_year", ""))
        ws.cell(row=row, column=6, value=node.get("_crossref_volume", ""))
        ws.cell(row=row, column=7, value=node.get("_crossref_pages", ""))
        ws.cell(row=row, column=8, value=node.get("_crossref_publisher", ""))
        ws.cell(row=row, column=9, value=node.get("_crossref_type", ""))
        ws.cell(row=row, column=10, value=node.get("_crossref_citations", ""))
        ws.cell(row=row, column=11, value=node.get("_crossref_abstract", ""))
        ws.cell(row=row, column=12, value=node.get("_crossref_url", ""))
        row += 1

    style_data(ws, row - 2, len(headers))
    auto_width(ws, len(headers))

    print(f"  Sheet 'Publications': {row - 2} rows")
    return ws


def write_pdb_sheet(wb, nodes, edges):
    """Sheet 4: PDB structures and their protein targets."""
    ws = wb.create_sheet("PDB Structures")

    headers = [
        "PDB ID", "Protein Target", "Protein ID",
        "Structure Title", "Method", "Resolution",
        "Chain", "Ligands"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    # Find HAS_STRUCTURE edges
    pdb_edges = [e for e in edges if e.get(":TYPE", "") == "HAS_STRUCTURE"]

    row = 2
    seen = set()
    for e in pdb_edges:
        src_id = e.get(":START_ID", "")
        tgt_id = e.get(":END_ID", "")
        key = (src_id, tgt_id)
        if key in seen:
            continue
        seen.add(key)

        src_node = nodes.get(src_id, {})
        tgt_node = nodes.get(tgt_id, {})

        pdb_id = tgt_node.get("name", tgt_id.replace("PDB:", "").replace("STRUCTURE:", ""))
        protein = src_node.get("name", src_id)

        ws.cell(row=row, column=1, value=pdb_id)
        ws.cell(row=row, column=2, value=protein)
        ws.cell(row=row, column=3, value=src_id)
        ws.cell(row=row, column=4, value=tgt_node.get("synonyms", ""))
        ws.cell(row=row, column=5, value=e.get("method", ""))
        ws.cell(row=row, column=6, value=e.get("resolution", ""))
        ws.cell(row=row, column=7, value=e.get("confidence", ""))
        ws.cell(row=row, column=8, value="")
        row += 1

    style_data(ws, row - 2, len(headers))
    auto_width(ws, len(headers))

    print(f"  Sheet 'PDB Structures': {row - 2} rows")
    return ws


def write_relations_sheet(wb, nodes, edges):
    """Sheet 5: All relations (full edge table)."""
    ws = wb.create_sheet("All Relations")

    headers = [
        "Source ID", "Source Name", "Source Type",
        "Target ID", "Target Name", "Target Type",
        "Relation Type", "Activity Type", "Activity Value",
        "Units", "pChEMBL", "Mechanism of Action", "Assay ID",
        "Confidence"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    row = 2
    for e in edges:
        src_id = e.get(":START_ID", "")
        tgt_id = e.get(":END_ID", "")
        src_node = nodes.get(src_id, {})
        tgt_node = nodes.get(tgt_id, {})

        ws.cell(row=row, column=1, value=src_id)
        ws.cell(row=row, column=2, value=src_node.get("name", ""))
        ws.cell(row=row, column=3, value=src_node.get(":LABEL", ""))
        ws.cell(row=row, column=4, value=tgt_id)
        ws.cell(row=row, column=5, value=tgt_node.get("name", ""))
        ws.cell(row=row, column=6, value=tgt_node.get(":LABEL", ""))
        ws.cell(row=row, column=7, value=e.get(":TYPE", ""))
        ws.cell(row=row, column=8, value=e.get("activity_type", ""))
        ws.cell(row=row, column=9, value=e.get("activity_value", ""))
        ws.cell(row=row, column=10, value=e.get("units", ""))
        ws.cell(row=row, column=11, value=e.get("pchembl_value", ""))
        ws.cell(row=row, column=12, value=e.get("mechanism_of_action", ""))
        ws.cell(row=row, column=13, value=e.get("assay_id", ""))
        ws.cell(row=row, column=14, value=e.get("confidence", ""))

        row += 1

    style_data(ws, row - 2, len(headers))
    auto_width(ws, len(headers))

    print(f"  Sheet 'All Relations': {row - 2} rows")
    return ws


# ── Summary Sheet ───────────────────────────────────────────────────────────

def write_summary_sheet(wb, compound_count, compound_enriched, inhibitor_count,
                         pub_count, pub_enriched, pdb_count, img_count, runtime):
    """Write a summary/README sheet."""
    ws = wb.create_sheet("Summary", 0)

    title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    section_font = Font(name="Calibri", bold=True, size=11, color="2F5496")
    normal_font = Font(name="Calibri", size=11)

    ws.cell(row=1, column=1, value="Chemical Data Extractor — Enrichment Report").font = title_font
    ws.merge_cells("A1:D1")

    ws.cell(row=3, column=1, value="Pipeline Statistics").font = section_font
    stats = [
        ("Total Compounds", str(compound_count)),
        ("Compounds with SMILES", f"{compound_enriched} ({compound_enriched/max(compound_count,1)*100:.0f}%)"),
        ("Inhibitor/Binder Activities", str(inhibitor_count)),
        ("Publications", str(pub_count)),
        ("Publications with CrossRef Metadata", str(pub_enriched)),
        ("PDB Structures", str(pdb_count)),
        ("2D Structure Images Downloaded", str(img_count)),
        ("Runtime", f"{runtime:.1f} seconds"),
    ]
    for i, (label, value) in enumerate(stats):
        ws.cell(row=4 + i, column=1, value=label).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=4 + i, column=2, value=value).font = normal_font

    ws.cell(row=14, column=1, value="Sheet Overview").font = section_font
    sheets = [
        "Summary — This page",
        "Compounds — All compounds with SMILES, properties, and 2D structure images",
        "Inhibitor Activities — Compound → Target inhibition/binding data with Ki/IC50 values",
        "Publications — Papers with DOIs, authors, abstracts, and citation counts",
        "PDB Structures — Protein-to-structure mappings with method and resolution",
        "All Relations — Complete edge table for all relationships",
    ]
    for i, s in enumerate(sheets):
        ws.cell(row=15 + i, column=1, value=s).font = normal_font

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50


# ── Main Pipeline ───────────────────────────────────────────────────────────

def run_enrichment(export_dir: str = "exports") -> None:
    """Run the full enrichment pipeline on a given export directory.
    
    Args:
        export_dir: Path to the directory containing nodes.csv and edges.csv
    """
    set_export_dir(export_dir)
    return _run_pipeline()


def _run_pipeline():
    start_time = time.time()
    print("=" * 60)
    print("  CHEMICAL DATA EXTRACTOR — ENRICHMENT PIPELINE")
    print("=" * 60)

    # 1. Load data
    print("\n📂 Loading export data...")
    nodes = load_nodes()
    edges = load_edges()
    print(f"  Loaded {len(nodes)} nodes and {len(edges)} edges")

    # 2. Enrich compounds
    print("\n🧪 Enriching compounds (PubChem)...")
    compound_nodes = enrich_compounds(nodes, edges)
    compound_enriched = sum(1 for n in compound_nodes.values() if n.get("_enriched_smiles"))

    # 3. Download structure images
    print("\n🖼️  Downloading 2D structure images...")
    img_count = download_structure_images(compound_nodes)

    # 4. Enrich publications
    print("\n📄 Enriching publications (CrossRef)...")
    pub_nodes = enrich_publications(nodes)
    pub_enriched = sum(1 for n in pub_nodes.values() if n.get("_crossref_title"))

    # 5. Count inhibitors
    inhibitor_edges = [e for e in edges if e.get(":TYPE", "") in ("INHIBITS", "BINDS", "ACTIVE_AGAINST")]
    pdb_edges = [e for e in edges if e.get(":TYPE", "") == "HAS_STRUCTURE"]

    # 6. Write Excel
    print("\n📊 Writing Excel workbook...")
    wb = openpyxl.Workbook()
    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_summary_sheet(wb, len(compound_nodes), compound_enriched,
                         len(inhibitor_edges), len(pub_nodes), pub_enriched,
                         len(pdb_edges), img_count, time.time() - start_time)
    write_compounds_sheet(wb, compound_nodes, nodes)
    write_inhibitors_sheet(wb, nodes, edges)
    write_publications_sheet(wb, pub_nodes)
    write_pdb_sheet(wb, nodes, edges)
    write_relations_sheet(wb, nodes, edges)

    wb.save(OUTPUT_XLSX)
    elapsed = time.time() - start_time

    print(f"\n✅ Enriched Excel saved: {OUTPUT_XLSX}")
    print(f"   Runtime: {elapsed:.1f} seconds")

    # Summary
    print(f"\n{'─' * 60}")
    print(f"  📊 ENRICHMENT SUMMARY")
    print(f"{'─' * 60}")
    print(f"  Compounds      : {len(compound_nodes):3d} ({compound_enriched} with SMILES)")
    print(f"  Inhibitor Edges: {len(inhibitor_edges):3d}")
    print(f"  Publications   : {len(pub_nodes):3d} ({pub_enriched} enriched)")
    print(f"  PDB Structures : {len(pdb_edges):3d}")
    print(f"  2D Images      : {img_count:3d}")
    print(f"{'─' * 60}")
    print(f"  Output file: {OUTPUT_XLSX}")
    print(f"{'─' * 60}")

    return compound_nodes, pub_nodes


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(description="Enrich chemical data exports with SMILES, properties, and publication metadata.")
    parser.add_argument("--export-dir", default="exports",
                        help="Directory containing nodes.csv and edges.csv (default: ./exports)")
    args = parser.parse_args()
    return run_enrichment(args.export_dir)


if __name__ == "__main__":
    main()
